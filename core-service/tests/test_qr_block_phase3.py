"""Phase 3 tests for Excel artifacts and private S3 downloads."""

from io import BytesIO
from types import SimpleNamespace
from unittest.mock import Mock
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.repositories.qr_activation_repository import (
    ProductItemRepository as ActivationItemRepository,
)
from app.repositories.qr_product_repository import (
    ProductItemRepository as BlockItemRepository,
)
from app.schemas.qr_product import QRBlockCreate
from app.services import storage_service
from app.services.qr_product_service import QRProductService, _build_excel


def make_service() -> QRProductService:
    service = QRProductService.__new__(QRProductService)
    service.db = Mock()
    service.product_repo = Mock()
    service.product_setting_repo = Mock()
    service.block_repo = Mock()
    service.item_repo = Mock()
    service.sku_repo = Mock()
    service.credit_service = Mock()
    service.key_service = None
    service.qr_shortener = Mock()
    return service


def test_excel_contains_signed_urls_and_optional_qr_images():
    data = _build_excel(
        [{"serial": "PRO-ABC12345", "primary_url": "https://qr.test/item"}],
        "dynamic",
        include_qr_images=True,
    )

    worksheet = load_workbook(BytesIO(data)).active
    assert [cell.value for cell in worksheet[1]] == [
        "QR URL",
        "Serial Number",
        "QR Image",
    ]
    assert worksheet.cell(2, 1).value == "https://qr.test/item"
    assert worksheet.cell(2, 2).value == "PRO-ABC12345"
    assert len(worksheet._images) == 1


def test_dual_excel_keeps_overt_and_covert_urls_separate():
    data = _build_excel(
        [{
            "serial": "PRO-ABC12345",
            "primary_url": "https://qr.test/overt",
            "overt_url": "https://qr.test/overt",
            "covert_url": "https://qr.test/covert",
        }],
        "dual",
    )

    worksheet = load_workbook(BytesIO(data)).active
    assert [cell.value for cell in worksheet[1]] == [
        "URL (Overt)",
        "URL (Covert)",
        "Serial Number",
    ]
    assert [worksheet.cell(2, column).value for column in range(1, 4)] == [
        "https://qr.test/overt",
        "https://qr.test/covert",
        "PRO-ABC12345",
    ]


@pytest.mark.parametrize(
    "activation_method, expected_active, expected_deactivated",
    [
        ("pre", True, False),
        ("post", False, True),
    ],
)
@pytest.mark.parametrize(
    "qr_type",
    ["dynamic", "static", "dual", "secure_code", "one_time", "post_activation"],
)
def test_all_qr_types_follow_product_activation_method(
    activation_method, expected_active, expected_deactivated, qr_type
):
    service = make_service()
    block = SimpleNamespace(
        id=uuid4(),
        product_id=uuid4(),
        sku_id=None,
        quantity=1,
        batch="ACTIVATION",
        qr_type=qr_type,
        sr_number_type="R8DAN",
        starting_serial=None,
        serial_prefix="PRO",
    )
    product = SimpleNamespace(
        brand_id=None,
        gtin="0123456789012",
        sr_number_type="R8DAN",
        activation_method=activation_method,
    )
    service.item_repo.get_existing_serials_global.return_value = set()

    service._generate_product_items(block, product, uuid4(), uuid4())

    item = service.item_repo.bulk_create.call_args.args[0][0]
    assert item["qr_active"] is expected_active
    assert item["qr_deactive"] is expected_deactivated
    assert item["qr_deactive_unit"] is expected_deactivated


def test_post_activation_uses_standard_gtin_long_url(monkeypatch):
    service = make_service()
    service.key_service = Mock()
    service.key_service.decrypt_private_key.return_value = object()
    service.key_service.sign_message.return_value = "signed-value"
    monkeypatch.setattr(
        "app.repositories.brand_repository.BrandRepository.get_by_id",
        Mock(return_value=SimpleNamespace(
            private_key_encrypted="encrypted",
            short_code="demo",
        )),
    )
    block = SimpleNamespace(
        id=uuid4(),
        product_id=uuid4(),
        sku_id=None,
        quantity=1,
        batch="POST-ACTIVATION",
        qr_type="post_activation",
        sr_number_type="R8DAN",
        starting_serial=None,
        serial_prefix="PRO",
    )
    product = SimpleNamespace(
        brand_id=uuid4(),
        gtin="0123456789012",
        sr_number_type="R8DAN",
        activation_method="post",
    )
    service.item_repo.get_existing_serials_global.return_value = set()
    service.qr_shortener.shorten.return_value = "https://bwqr.me/post-item"

    service._generate_product_items(block, product, uuid4(), uuid4())

    item = service.item_repo.bulk_create.call_args.args[0][0]
    assert item["token_id"] == "https://bwqr.me/post-item"
    assert "/g/0123456789012/s/" in item["extra_data"]["long_url"]
    assert "qsealpost" not in item["extra_data"]["long_url"]


def test_block_detail_activation_summary_is_tenant_scoped():
    service = make_service()
    organization_id = uuid4()
    block_id = uuid4()
    block = SimpleNamespace(id=block_id)
    service.block_repo.get_by_id.return_value = block
    service.item_repo.get_activation_summary.return_value = (10, 4)

    result = service.get_block_detail(block_id, organization_id)

    assert result.activation_status == "partially_activated"
    assert result.activated_count == 4
    assert result.deactivated_count == 6
    service.item_repo.get_activation_summary.assert_called_once_with(
        block_id, organization_id
    )


def test_activation_summary_query_is_organization_scoped():
    organization_id = uuid4()
    block_id = uuid4()
    db = Mock()
    query = db.query.return_value
    query.filter.return_value = query
    query.one.return_value = (2, 1)

    assert BlockItemRepository(db).get_activation_summary(
        block_id, organization_id
    ) == (2, 1)
    filters = query.filter.call_args.args
    assert any(
        "product_items.organization_id" in str(expression)
        for expression in filters
    )
    assert any(
        "product_items.block_id" in str(expression)
        for expression in filters
    )


def test_activation_updates_all_item_state_flags():
    repository = ActivationItemRepository(Mock())
    item = SimpleNamespace(
        qr_active=False,
        qr_deactive=True,
        qr_deactive_unit=True,
    )

    repository.update_qr_status(item)

    assert item.qr_active is True
    assert item.qr_deactive is False
    assert item.qr_deactive_unit is False


def test_s3_artifact_is_private_and_tenant_scoped(monkeypatch):
    organization_id = uuid4()
    product_id = uuid4()
    block_id = uuid4()
    client = Mock()
    monkeypatch.setattr(storage_service.settings, "aws_s3_bucket", "private-bucket")
    monkeypatch.setattr(storage_service, "_get_s3_client", lambda: client)

    object_key = storage_service.build_qr_artifact_key(
        organization_id, product_id, block_id
    )
    storage_service.store_qr_artifact(b"workbook", object_key, "codes.xlsx")

    assert str(organization_id) in object_key
    assert str(product_id) in object_key
    assert str(block_id) in object_key
    client.put_object.assert_called_once_with(
        Bucket="private-bucket",
        Key=object_key,
        Body=b"workbook",
        ContentType=storage_service.QR_ARTIFACT_CONTENT_TYPE,
        ContentDisposition='attachment; filename="codes.xlsx"',
        ServerSideEncryption="AES256",
    )


def test_download_uses_scoped_block_lookup_and_presigned_artifact(monkeypatch):
    service = make_service()
    organization_id = uuid4()
    block_id = uuid4()
    object_key = f"qseal/organizations/{organization_id}/blocks/{block_id}.xlsx"
    block = SimpleNamespace(
        id=block_id,
        status="completed",
        artifact_object_key=object_key,
        download_url=None,
    )
    service.block_repo.get_by_id.return_value = block
    signed = ("https://signed.test/artifact", Mock())
    monkeypatch.setattr(
        storage_service,
        "get_qr_artifact_signed_url",
        Mock(return_value=signed),
    )

    result = service.get_block_download_url(block_id, organization_id)

    assert result == signed
    service.block_repo.get_by_id.assert_called_once_with(block_id, organization_id)
    storage_service.get_qr_artifact_signed_url.assert_called_once_with(object_key)


def test_artifact_failure_marks_block_failed_without_deducting_credits():
    service = make_service()
    organization_id = uuid4()
    product = SimpleNamespace(
        id=uuid4(), qr_type="dynamic", sr_number_type="R8DAN", serial_prefix="PRO"
    )
    block = SimpleNamespace(
        id=uuid4(), status="pending", task_status="pending",
        error_code=None, error_message=None,
    )
    service.product_repo.get_by_id.return_value = product
    service.block_repo.batch_exists.return_value = False
    service.block_repo.create.return_value = block
    service.block_repo.get_by_id.return_value = block
    service._generate_product_items = Mock(return_value=[])
    service._store_block_artifact = Mock(side_effect=RuntimeError("S3 unavailable"))

    with pytest.raises(RuntimeError, match="S3 unavailable"):
        service.generate_block(
            product.id,
            QRBlockCreate(batch="PHASE3", quantity=1),
            organization_id,
            uuid4(),
        )

    assert block.status == "failed"
    service.credit_service.consume_reserved_credits.assert_not_called()
    service.credit_service.release_reserved_credits.assert_called_once_with(
        organization_id,
        block.id,
    )
    service.item_repo.soft_delete_by_block.assert_called_once_with(
        block.id, organization_id
    )


def test_credit_failure_cleans_up_uploaded_artifact(monkeypatch):
    service = make_service()
    organization_id = uuid4()
    product = SimpleNamespace(
        id=uuid4(), qr_type="dynamic", sr_number_type="R8DAN", serial_prefix="PRO"
    )
    block = SimpleNamespace(
        id=uuid4(), status="pending", task_status="pending",
        error_code=None, error_message=None,
    )
    service.product_repo.get_by_id.return_value = product
    service.block_repo.batch_exists.return_value = False
    service.block_repo.create.return_value = block
    service.block_repo.get_by_id.return_value = block
    service._generate_product_items = Mock(return_value=[])
    service._store_block_artifact = Mock(return_value="tenant/artifact.xlsx")
    service.credit_service.consume_reserved_credits.side_effect = RuntimeError(
        "credit transaction failed"
    )
    delete_artifact = Mock()
    monkeypatch.setattr(storage_service, "delete_qr_artifact", delete_artifact)

    with pytest.raises(RuntimeError, match="credit transaction failed"):
        service.generate_block(
            product.id,
            QRBlockCreate(batch="PHASE3-CLEANUP", quantity=1),
            organization_id,
            uuid4(),
        )

    delete_artifact.assert_called_once_with("tenant/artifact.xlsx")
    assert block.status == "failed"


def test_artifact_presign_failure_returns_service_unavailable(monkeypatch):
    service = make_service()
    organization_id = uuid4()
    block_id = uuid4()
    service.block_repo.get_by_id.return_value = SimpleNamespace(
        id=block_id,
        status="completed",
        artifact_object_key="tenant/artifact.xlsx",
        download_url=None,
    )
    monkeypatch.setattr(
        storage_service,
        "get_qr_artifact_signed_url",
        Mock(side_effect=RuntimeError("AWS unavailable")),
    )

    with pytest.raises(HTTPException) as exc_info:
        service.get_block_download_url(block_id, organization_id)

    assert exc_info.value.status_code == 503
