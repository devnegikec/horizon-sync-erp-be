"""Tests for Payment Export Service"""

import io
from datetime import datetime
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from app.services.payment_export_service import PaymentExportService


class MockReconciliationService:
    """Mock reconciliation service for testing"""
    
    def generate_report(
        self,
        organization_id,
        date_from=None,
        date_to=None,
        party_id=None,
        payment_mode=None,
        status=None,
    ):
        """Generate mock report data"""
        return {
            "summary": {
                "total_payments_received": "15000.00",
                "total_allocated": "12000.00",
                "total_unallocated": "3000.00",
                "payment_count": 5,
                "unallocated_payment_count": 2,
            },
            "payments_by_status": {
                "Confirmed": {
                    "count": 3,
                    "total_amount": "10000.00",
                    "payments": [],
                },
                "Draft": {
                    "count": 2,
                    "total_amount": "5000.00",
                    "payments": [],
                },
            },
            "payments_by_mode": {
                "Cash": {
                    "count": 2,
                    "total_amount": "3000.00",
                    "payments": [],
                },
                "Bank_Transfer": {
                    "count": 2,
                    "total_amount": "10000.00",
                    "payments": [],
                },
                "Check": {
                    "count": 1,
                    "total_amount": "2000.00",
                    "payments": [],
                },
            },
            "payments": [
                {
                    "id": str(uuid4()),
                    "payment_type": "Customer_Payment",
                    "party_id": str(uuid4()),
                    "amount": "5000.00",
                    "currency_code": "USD",
                    "payment_date": "2024-01-15T10:00:00",
                    "payment_mode": "Bank_Transfer",
                    "reference_no": "TXN123456",
                    "status": "Confirmed",
                    "receipt_number": "RCP-2024-001",
                    "unallocated_amount": "1000.00",
                    "allocated_invoices": [
                        {
                            "invoice_id": str(uuid4()),
                            "invoice_number": "INV-001",
                            "allocated_amount": "4000.00",
                        }
                    ],
                    "allocation_count": 1,
                    "created_at": "2024-01-15T10:00:00",
                },
                {
                    "id": str(uuid4()),
                    "payment_type": "Customer_Payment",
                    "party_id": str(uuid4()),
                    "amount": "3000.00",
                    "currency_code": "USD",
                    "payment_date": "2024-01-16T11:00:00",
                    "payment_mode": "Cash",
                    "reference_no": None,
                    "status": "Confirmed",
                    "receipt_number": "RCP-2024-002",
                    "unallocated_amount": "0.00",
                    "allocated_invoices": [
                        {
                            "invoice_id": str(uuid4()),
                            "invoice_number": "INV-002",
                            "allocated_amount": "3000.00",
                        }
                    ],
                    "allocation_count": 1,
                    "created_at": "2024-01-16T11:00:00",
                },
                {
                    "id": str(uuid4()),
                    "payment_type": "Supplier_Payment",
                    "party_id": str(uuid4()),
                    "amount": "7000.00",
                    "currency_code": "USD",
                    "payment_date": "2024-01-17T12:00:00",
                    "payment_mode": "Bank_Transfer",
                    "reference_no": "TXN789012",
                    "status": "Confirmed",
                    "receipt_number": "RCP-2024-003",
                    "unallocated_amount": "2000.00",
                    "allocated_invoices": [
                        {
                            "invoice_id": str(uuid4()),
                            "invoice_number": "PINV-001",
                            "allocated_amount": "5000.00",
                        }
                    ],
                    "allocation_count": 1,
                    "created_at": "2024-01-17T12:00:00",
                },
            ],
            "unallocated_payments": [
                {
                    "id": str(uuid4()),
                    "payment_type": "Customer_Payment",
                    "party_id": str(uuid4()),
                    "amount": "5000.00",
                    "currency_code": "USD",
                    "payment_date": "2024-01-15T10:00:00",
                    "payment_mode": "Bank_Transfer",
                    "reference_no": "TXN123456",
                    "status": "Confirmed",
                    "receipt_number": "RCP-2024-001",
                    "unallocated_amount": "1000.00",
                    "allocated_invoices": [],
                    "allocation_count": 1,
                    "created_at": "2024-01-15T10:00:00",
                },
                {
                    "id": str(uuid4()),
                    "payment_type": "Supplier_Payment",
                    "party_id": str(uuid4()),
                    "amount": "7000.00",
                    "currency_code": "USD",
                    "payment_date": "2024-01-17T12:00:00",
                    "payment_mode": "Bank_Transfer",
                    "reference_no": "TXN789012",
                    "status": "Confirmed",
                    "receipt_number": "RCP-2024-003",
                    "unallocated_amount": "2000.00",
                    "allocated_invoices": [],
                    "allocation_count": 1,
                    "created_at": "2024-01-17T12:00:00",
                },
            ],
            "filters": {
                "date_from": "2024-01-01" if date_from else None,
                "date_to": "2024-01-31" if date_to else None,
                "party_id": str(party_id) if party_id else None,
                "payment_mode": payment_mode,
                "status": status,
            },
        }


@pytest.fixture
def export_service():
    """Create export service with mock reconciliation service"""
    mock_reconciliation = MockReconciliationService()
    return PaymentExportService(mock_reconciliation)


@pytest.fixture
def organization_id():
    """Generate test organization ID"""
    return uuid4()


class TestExportToExcel:
    """Tests for Excel export functionality"""
    
    def test_export_to_excel_creates_valid_workbook(self, export_service, organization_id):
        """Test that Excel export creates a valid workbook"""
        # Export to Excel
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Organization"
        )
        
        # Verify data is bytes
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 0
        
        # Load workbook
        wb = load_workbook(io.BytesIO(excel_data))
        
        # Verify sheets exist
        assert "Summary" in wb.sheetnames
        assert "Payment Details" in wb.sheetnames
        assert "Unallocated Payments" in wb.sheetnames
    
    def test_summary_sheet_contains_organization_name(self, export_service, organization_id):
        """Test that summary sheet includes organization branding"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="ACME Corporation"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Summary"]
        
        # Check organization name in first cell
        assert ws.cell(row=1, column=1).value == "ACME Corporation"
    
    def test_summary_sheet_contains_report_title(self, export_service, organization_id):
        """Test that summary sheet includes report title"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Summary"]
        
        # Check report title
        assert ws.cell(row=2, column=1).value == "Payment Reconciliation Report"
    
    def test_summary_sheet_contains_totals(self, export_service, organization_id):
        """Test that summary sheet includes summary totals"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Summary"]
        
        # Find summary section (should contain these labels)
        found_labels = []
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=1):
            cell_value = row[0].value
            if cell_value:
                found_labels.append(str(cell_value))
        
        assert "Total Payments Received" in found_labels
        assert "Total Allocated" in found_labels
        assert "Total Unallocated" in found_labels
    
    def test_summary_sheet_contains_status_breakdown(self, export_service, organization_id):
        """Test that summary sheet includes payments by status"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Summary"]
        
        # Find "Payments by Status" section
        found_section = False
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1):
            if row[0].value == "Payments by Status":
                found_section = True
                break
        
        assert found_section, "Payments by Status section not found"
    
    def test_summary_sheet_contains_mode_breakdown(self, export_service, organization_id):
        """Test that summary sheet includes payments by mode"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Summary"]
        
        # Find "Payments by Mode" section
        found_section = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1):
            if row[0].value == "Payments by Mode":
                found_section = True
                break
        
        assert found_section, "Payments by Mode section not found"
    
    def test_payment_details_sheet_has_headers(self, export_service, organization_id):
        """Test that payment details sheet has proper headers"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Payment Details"]
        
        # Check headers in first row
        expected_headers = [
            "Receipt Number",
            "Payment Date",
            "Payment Type",
            "Amount",
            "Currency",
            "Payment Mode",
            "Reference No",
            "Status",
            "Allocated Amount",
            "Unallocated Amount",
            "Allocation Count",
        ]
        
        for col_num, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col_num).value
            assert actual_header == expected_header
    
    def test_payment_details_sheet_contains_data(self, export_service, organization_id):
        """Test that payment details sheet contains payment data"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Payment Details"]
        
        # Check that we have data rows (more than just header)
        assert ws.max_row > 1
        
        # Check first data row has receipt number
        receipt_number = ws.cell(row=2, column=1).value
        assert receipt_number is not None
        assert "RCP-" in receipt_number
    
    def test_unallocated_sheet_has_headers(self, export_service, organization_id):
        """Test that unallocated payments sheet has proper headers"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Unallocated Payments"]
        
        # Check headers in first row
        expected_headers = [
            "Receipt Number",
            "Payment Date",
            "Payment Type",
            "Total Amount",
            "Currency",
            "Payment Mode",
            "Status",
            "Unallocated Amount",
        ]
        
        for col_num, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col_num).value
            assert actual_header == expected_header
    
    def test_unallocated_sheet_contains_only_unallocated_payments(self, export_service, organization_id):
        """Test that unallocated sheet only shows payments with unallocated amounts"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Unallocated Payments"]
        
        # Check that we have data rows
        assert ws.max_row > 1
        
        # Check that all data rows have unallocated amount > 0
        for row_num in range(2, ws.max_row + 1):
            unallocated = ws.cell(row=row_num, column=8).value
            if unallocated is not None:
                assert float(unallocated) > 0
    
    def test_export_with_date_filters(self, export_service, organization_id):
        """Test Excel export with date filters"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 1, 31)
        
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            date_from=date_from,
            date_to=date_to,
            organization_name="Test Org"
        )
        
        # Verify export succeeds
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 0
    
    def test_export_with_all_filters(self, export_service, organization_id):
        """Test Excel export with all filters applied"""
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            date_from=datetime(2024, 1, 1),
            date_to=datetime(2024, 1, 31),
            party_id=uuid4(),
            payment_mode="Bank_Transfer",
            status="Confirmed",
            organization_name="Test Org"
        )
        
        # Verify export succeeds
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 0


class TestExportToPDF:
    """Tests for PDF export functionality"""
    
    def test_export_to_pdf_creates_valid_pdf(self, export_service, organization_id):
        """Test that PDF export creates a valid PDF"""
        # Export to PDF
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Organization"
        )
        
        # Verify data is bytes
        assert isinstance(pdf_data, bytes)
        assert len(pdf_data) > 0
        
        # Verify it's a valid PDF
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        assert len(pdf_reader.pages) > 0
    
    def test_pdf_contains_organization_name(self, export_service, organization_id):
        """Test that PDF includes organization branding"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="ACME Corporation"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for organization name
        assert "ACME Corporation" in first_page_text
    
    def test_pdf_contains_report_title(self, export_service, organization_id):
        """Test that PDF includes report title"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for report title
        assert "Payment Reconciliation Report" in first_page_text
    
    def test_pdf_contains_summary_section(self, export_service, organization_id):
        """Test that PDF includes summary section"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for summary section
        assert "Summary" in first_page_text
        assert "Total Payments Received" in first_page_text
        assert "Total Allocated" in first_page_text
    
    def test_pdf_contains_status_breakdown(self, export_service, organization_id):
        """Test that PDF includes payments by status"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for status breakdown
        assert "Payments by Status" in first_page_text
    
    def test_pdf_contains_mode_breakdown(self, export_service, organization_id):
        """Test that PDF includes payments by mode"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for mode breakdown
        assert "Payments by Mode" in first_page_text
    
    def test_pdf_contains_payment_details(self, export_service, organization_id):
        """Test that PDF includes payment details"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        all_text = ""
        for page in pdf_reader.pages:
            all_text += page.extract_text()
        
        # Check for payment data (receipt numbers)
        assert "RCP-2024-" in all_text
    
    def test_pdf_with_date_filters_shows_period(self, export_service, organization_id):
        """Test that PDF shows date period when filters applied"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 1, 31)
        
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            date_from=date_from,
            date_to=date_to,
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for period information
        assert "Period:" in first_page_text or "2024-01-01" in first_page_text
    
    def test_pdf_with_filters_shows_filter_info(self, export_service, organization_id):
        """Test that PDF shows filter information"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            payment_mode="Bank_Transfer",
            status="Confirmed",
            organization_name="Test Org"
        )
        
        # Read PDF text
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        first_page_text = pdf_reader.pages[0].extract_text()
        
        # Check for filter information
        assert "Mode:" in first_page_text or "Bank_Transfer" in first_page_text
        assert "Status:" in first_page_text or "Confirmed" in first_page_text
    
    def test_export_with_all_filters_to_pdf(self, export_service, organization_id):
        """Test PDF export with all filters applied"""
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            date_from=datetime(2024, 1, 1),
            date_to=datetime(2024, 1, 31),
            party_id=uuid4(),
            payment_mode="Cash",
            status="Draft",
            organization_name="Test Org"
        )
        
        # Verify export succeeds
        assert isinstance(pdf_data, bytes)
        assert len(pdf_data) > 0
        
        # Verify it's a valid PDF
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        assert len(pdf_reader.pages) > 0


class TestExportServiceIntegration:
    """Integration tests for export service"""
    
    def test_both_exports_use_same_data(self, export_service, organization_id):
        """Test that Excel and PDF exports use the same underlying data"""
        # Export to both formats
        excel_data = export_service.export_to_excel(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        pdf_data = export_service.export_to_pdf(
            organization_id=organization_id,
            organization_name="Test Org"
        )
        
        # Both should succeed
        assert len(excel_data) > 0
        assert len(pdf_data) > 0
        
        # Verify Excel contains expected data
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb["Summary"]
        
        # Find total payments value in Excel
        excel_total = None
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
            if row[0].value == "Total Payments Received":
                excel_total = row[1].value
                break
        
        # Verify PDF contains same total
        pdf_reader = PdfReader(io.BytesIO(pdf_data))
        pdf_text = pdf_reader.pages[0].extract_text()
        
        # Both should reference the same total amount
        assert excel_total is not None
        assert "15000" in str(excel_total) or "15,000" in str(excel_total)
        assert "15000" in pdf_text or "15,000" in pdf_text
