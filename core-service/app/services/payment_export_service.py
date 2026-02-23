"""Export service for Payment Reconciliation Reports"""

import io
import logging
from datetime import datetime
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from app.services.reconciliation_report_service import ReconciliationReportService

logger = logging.getLogger(__name__)


class PaymentExportService:
    """
    Service for exporting Payment Reconciliation Report data in various formats.
    
    Supports:
    - XLSX export (Excel) with organization branding
    - PDF export with organization branding
    """
    
    def __init__(self, reconciliation_service: ReconciliationReportService):
        """
        Initialize payment export service
        
        Args:
            reconciliation_service: Reconciliation report service for generating data
        """
        self.reconciliation_service = reconciliation_service
    
    def export_to_excel(
        self,
        organization_id: UUID,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        party_id: Optional[UUID] = None,
        payment_mode: Optional[str] = None,
        status: Optional[str] = None,
        organization_name: str = "Organization",
    ) -> bytes:
        """
        Export Payment Reconciliation Report to Excel format.
        
        Args:
            organization_id: Organization UUID
            date_from: Start date for payment date range (optional)
            date_to: End date for payment date range (optional)
            party_id: Filter by customer/supplier ID (optional)
            payment_mode: Filter by payment mode (optional)
            status: Filter by payment status (optional)
            organization_name: Organization name for branding
            
        Returns:
            XLSX data as bytes
        """
        # Generate report data
        report = self.reconciliation_service.generate_report(
            organization_id=organization_id,
            date_from=date_from,
            date_to=date_to,
            party_id=party_id,
            payment_mode=payment_mode,
            status=status,
        )
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet and create named sheets
        wb.remove(wb.active)
        
        # Create Summary sheet
        self._create_summary_sheet(wb, report, organization_name)
        
        # Create Payments Detail sheet
        self._create_payments_detail_sheet(wb, report)
        
        # Create Unallocated Payments sheet
        self._create_unallocated_sheet(wb, report)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        xlsx_data = output.getvalue()
        output.close()
        
        return xlsx_data
    
    def _create_summary_sheet(self, wb: Workbook, report: dict, organization_name: str):
        """Create summary sheet with report overview"""
        ws = wb.create_sheet("Summary")
        
        # Define styles
        title_font = Font(bold=True, size=14, color="366092")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Organization branding
        ws.cell(row=1, column=1, value=organization_name).font = title_font
        ws.cell(row=2, column=1, value="Payment Reconciliation Report").font = Font(bold=True, size=12)
        
        # Report period
        filters = report["filters"]
        period_text = "Period: "
        if filters["date_from"] and filters["date_to"]:
            period_text += f"{filters['date_from']} to {filters['date_to']}"
        elif filters["date_from"]:
            period_text += f"From {filters['date_from']}"
        elif filters["date_to"]:
            period_text += f"Until {filters['date_to']}"
        else:
            period_text += "All Time"
        
        ws.cell(row=3, column=1, value=period_text)
        ws.cell(row=4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Summary totals
        row = 6
        ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=11)
        row += 1
        
        summary = report["summary"]
        summary_data = [
            ["Total Payments Received", summary["total_payments_received"]],
            ["Total Allocated", summary["total_allocated"]],
            ["Total Unallocated", summary["total_unallocated"]],
            ["Payment Count", summary["payment_count"]],
            ["Unallocated Payment Count", summary["unallocated_payment_count"]],
        ]
        
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Payments by Status
        row += 2
        ws.cell(row=row, column=1, value="Payments by Status").font = Font(bold=True, size=11)
        row += 1
        
        # Header
        ws.cell(row=row, column=1, value="Status").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Count").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3, value="Total Amount").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        row += 1
        
        for status, data in report["payments_by_status"].items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=float(data["total_amount"]))
            row += 1
        
        # Payments by Mode
        row += 2
        ws.cell(row=row, column=1, value="Payments by Mode").font = Font(bold=True, size=11)
        row += 1
        
        # Header
        ws.cell(row=row, column=1, value="Payment Mode").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Count").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3, value="Total Amount").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        row += 1
        
        for mode, data in report["payments_by_mode"].items():
            ws.cell(row=row, column=1, value=mode)
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=float(data["total_amount"]))
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    def _create_payments_detail_sheet(self, wb: Workbook, report: dict):
        """Create detailed payments sheet"""
        ws = wb.create_sheet("Payment Details")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Receipt Number",
            "Payment Date",
            "Payment Type",
            "Amount",
            "Currency",
            "Payment Mode",
            "Reference No",
            "Status",
            "Allocated Amount",
            "Unallocated Amount",
            "Allocation Count",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_num, payment in enumerate(report["payments"], 2):
            ws.cell(row=row_num, column=1, value=payment.get("receipt_number", "")).border = thin_border
            ws.cell(row=row_num, column=2, value=payment["payment_date"]).border = thin_border
            ws.cell(row=row_num, column=3, value=payment["payment_type"]).border = thin_border
            ws.cell(row=row_num, column=4, value=float(payment["amount"])).border = thin_border
            ws.cell(row=row_num, column=5, value=payment["currency_code"]).border = thin_border
            ws.cell(row=row_num, column=6, value=payment["payment_mode"]).border = thin_border
            ws.cell(row=row_num, column=7, value=payment.get("reference_no", "")).border = thin_border
            ws.cell(row=row_num, column=8, value=payment["status"]).border = thin_border
            
            # Calculate allocated amount
            allocated = float(payment["amount"]) - float(payment["unallocated_amount"])
            ws.cell(row=row_num, column=9, value=allocated).border = thin_border
            ws.cell(row=row_num, column=10, value=float(payment["unallocated_amount"])).border = thin_border
            ws.cell(row=row_num, column=11, value=payment["allocation_count"]).border = thin_border
        
        # Auto-adjust column widths
        column_widths = [18, 12, 15, 12, 10, 15, 15, 12, 15, 15, 12]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
    
    def _create_unallocated_sheet(self, wb: Workbook, report: dict):
        """Create unallocated payments sheet"""
        ws = wb.create_sheet("Unallocated Payments")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Receipt Number",
            "Payment Date",
            "Payment Type",
            "Total Amount",
            "Currency",
            "Payment Mode",
            "Status",
            "Unallocated Amount",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_num, payment in enumerate(report["unallocated_payments"], 2):
            ws.cell(row=row_num, column=1, value=payment.get("receipt_number", "")).border = thin_border
            ws.cell(row=row_num, column=2, value=payment["payment_date"]).border = thin_border
            ws.cell(row=row_num, column=3, value=payment["payment_type"]).border = thin_border
            ws.cell(row=row_num, column=4, value=float(payment["amount"])).border = thin_border
            ws.cell(row=row_num, column=5, value=payment["currency_code"]).border = thin_border
            ws.cell(row=row_num, column=6, value=payment["payment_mode"]).border = thin_border
            ws.cell(row=row_num, column=7, value=payment["status"]).border = thin_border
            ws.cell(row=row_num, column=8, value=float(payment["unallocated_amount"])).border = thin_border
        
        # Auto-adjust column widths
        column_widths = [18, 12, 15, 12, 10, 15, 12, 15]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
    
    def export_to_pdf(
        self,
        organization_id: UUID,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        party_id: Optional[UUID] = None,
        payment_mode: Optional[str] = None,
        status: Optional[str] = None,
        organization_name: str = "Organization",
    ) -> bytes:
        """
        Export Payment Reconciliation Report to PDF format.
        
        Args:
            organization_id: Organization UUID
            date_from: Start date for payment date range (optional)
            date_to: End date for payment date range (optional)
            party_id: Filter by customer/supplier ID (optional)
            payment_mode: Filter by payment mode (optional)
            status: Filter by payment status (optional)
            organization_name: Organization name for branding
            
        Returns:
            PDF data as bytes
        """
        # Generate report data
        report = self.reconciliation_service.generate_report(
            organization_id=organization_id,
            date_from=date_from,
            date_to=date_to,
            party_id=party_id,
            payment_mode=payment_mode,
            status=status,
        )
        
        # Create PDF in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        # Container for PDF elements
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        
        # Organization branding
        org_style = ParagraphStyle(
            'OrgName',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
        )
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )
        
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20,
        )
        
        # Add organization name
        org_name = Paragraph(organization_name, org_style)
        elements.append(org_name)
        
        # Add title
        title = Paragraph("Payment Reconciliation Report", title_style)
        elements.append(title)
        
        # Add report metadata
        filters = report["filters"]
        metadata_parts = []
        
        if filters["date_from"] and filters["date_to"]:
            metadata_parts.append(f"Period: {filters['date_from']} to {filters['date_to']}")
        elif filters["date_from"]:
            metadata_parts.append(f"From: {filters['date_from']}")
        elif filters["date_to"]:
            metadata_parts.append(f"Until: {filters['date_to']}")
        
        if filters["payment_mode"]:
            metadata_parts.append(f"Mode: {filters['payment_mode']}")
        if filters["status"]:
            metadata_parts.append(f"Status: {filters['status']}")
        
        metadata_parts.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        metadata_text = " | ".join(metadata_parts)
        metadata = Paragraph(metadata_text, metadata_style)
        elements.append(metadata)
        
        # Summary section
        self._add_summary_section(elements, report, styles)
        
        elements.append(Spacer(1, 20))
        
        # Payments by Status section
        self._add_status_breakdown_section(elements, report, styles)
        
        elements.append(Spacer(1, 20))
        
        # Payments by Mode section
        self._add_mode_breakdown_section(elements, report, styles)
        
        elements.append(Spacer(1, 20))
        
        # Payment details table
        self._add_payment_details_table(elements, report)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        pdf_data = output.getvalue()
        output.close()
        
        return pdf_data
    
    def _add_summary_section(self, elements: list, report: dict, styles):
        """Add summary section to PDF"""
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#366092'),
            spaceAfter=10,
            fontName='Helvetica-Bold',
        )
        
        section_title = Paragraph("Summary", section_style)
        elements.append(section_title)
        
        summary = report["summary"]
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Payments Received", f"${float(summary['total_payments_received']):,.2f}"],
            ["Total Allocated", f"${float(summary['total_allocated']):,.2f}"],
            ["Total Unallocated", f"${float(summary['total_unallocated']):,.2f}"],
            ["Payment Count", str(summary["payment_count"])],
            ["Unallocated Payment Count", str(summary["unallocated_payment_count"])],
        ]
        
        table = Table(summary_data, colWidths=[3.5*inch, 2.5*inch])
        
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
    
    def _add_status_breakdown_section(self, elements: list, report: dict, styles):
        """Add payments by status breakdown to PDF"""
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#366092'),
            spaceAfter=10,
            fontName='Helvetica-Bold',
        )
        
        section_title = Paragraph("Payments by Status", section_style)
        elements.append(section_title)
        
        status_data = [["Status", "Count", "Total Amount"]]
        
        for status, data in report["payments_by_status"].items():
            status_data.append([
                status,
                str(data["count"]),
                f"${float(data['total_amount']):,.2f}"
            ])
        
        table = Table(status_data, colWidths=[2*inch, 2*inch, 2*inch])
        
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
    
    def _add_mode_breakdown_section(self, elements: list, report: dict, styles):
        """Add payments by mode breakdown to PDF"""
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#366092'),
            spaceAfter=10,
            fontName='Helvetica-Bold',
        )
        
        section_title = Paragraph("Payments by Mode", section_style)
        elements.append(section_title)
        
        mode_data = [["Payment Mode", "Count", "Total Amount"]]
        
        for mode, data in report["payments_by_mode"].items():
            mode_data.append([
                mode,
                str(data["count"]),
                f"${float(data['total_amount']):,.2f}"
            ])
        
        table = Table(mode_data, colWidths=[2*inch, 2*inch, 2*inch])
        
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
    
    def _add_payment_details_table(self, elements: list, report: dict):
        """Add detailed payment table to PDF"""
        # Limit to first 50 payments to avoid PDF size issues
        payments = report["payments"][:50]
        
        if not payments:
            return
        
        # Prepare table data
        table_data = [
            ["Receipt", "Date", "Type", "Amount", "Mode", "Status", "Unalloc."]
        ]
        
        for payment in payments:
            table_data.append([
                payment.get("receipt_number", "")[:12],
                payment["payment_date"][:10],
                payment["payment_type"][:8],
                f"${float(payment['amount']):,.0f}",
                payment["payment_mode"][:8],
                payment["status"][:8],
                f"${float(payment['unallocated_amount']):,.0f}",
            ])
        
        # Create table
        table = Table(table_data, colWidths=[
            1.0*inch,  # Receipt
            0.9*inch,  # Date
            0.8*inch,  # Type
            0.8*inch,  # Amount
            0.8*inch,  # Mode
            0.7*inch,  # Status
            0.8*inch,  # Unallocated
        ])
        
        # Style the table
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Amount
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Unallocated
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add note if payments were truncated
        if len(report["payments"]) > 50:
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
            note_style = ParagraphStyle(
                'Note',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER,
                spaceAfter=10,
            )
            note = Paragraph(
                f"Note: Showing first 50 of {len(report['payments'])} payments. "
                "Export to Excel for complete data.",
                note_style
            )
            elements.append(Spacer(1, 10))
            elements.append(note)
