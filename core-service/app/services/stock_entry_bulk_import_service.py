"""Bulk import service for stock entries — CSV and XLSX, using human-readable codes."""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from uuid import UUID

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.item import Item
from app.models.warehouse import Warehouse
from app.services.stock_entry_service import StockEntryService
from app.schemas.stock_entry import StockEntryCreate, StockEntryItemCreate

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column names (exactly as they appear in the CSV/XLSX header row)
# ---------------------------------------------------------------------------

TEMPLATE_HEADERS = [
    "Stock Entry Type",
    "Posting Date",
    "Posting Time",
    "From Warehouse Code",
    "To Warehouse Code",
    "Item Code",
    "Description",
    "Remarks",
    "Quantity",
    "UOM",
    "Basic Rate",
    "Valuation Rate",
    "Batch Number",
]

# Normalised (lowercase + underscores) versions used internally
_NORM_HEADERS = [h.lower().replace(" ", "_") for h in TEMPLATE_HEADERS]

REQUIRED_COLUMNS = {"stock_entry_type", "posting_date", "item_code", "quantity", "uom"}

VALID_ENTRY_TYPES = {
    "material_receipt",
    "material_issue",
    "material_transfer",
    "manufacture",
    "repack",
    "send_to_subcontractor",
}

# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------


@dataclass
class RowError:
    row: int
    field: str
    message: str


@dataclass
class BulkImportResult:
    total_rows: int = 0
    created: int = 0
    failed: int = 0
    errors: list[RowError] = field(default_factory=list)

    def add_error(self, row: int, field_name: str, message: str) -> None:
        self.errors.append(RowError(row=row, field=field_name, message=message))
        self.failed += 1


# ---------------------------------------------------------------------------
# Code → UUID lookup (cached per import run)
# ---------------------------------------------------------------------------


class _CodeCache:
    """Lazy-loads item and warehouse codes for the organisation."""

    def __init__(self, db: Session, organization_id: UUID):
        self._db = db
        self._org = organization_id
        self._items: dict[str, UUID] | None = None
        self._warehouses: dict[str, UUID] | None = None

    def _load_items(self) -> None:
        rows = (
            self._db.query(Item.item_code, Item.id)
            .filter(Item.organization_id == self._org)
            .all()
        )
        self._items = {r.item_code.strip().lower(): r.id for r in rows}

    def _load_warehouses(self) -> None:
        rows = (
            self._db.query(Warehouse.code, Warehouse.id)
            .filter(Warehouse.organization_id == self._org)
            .all()
        )
        self._warehouses = {r.code.strip().lower(): r.id for r in rows}

    def item_id(self, code: str) -> UUID | None:
        if self._items is None:
            self._load_items()
        return self._items.get(code.strip().lower())  # type: ignore[union-attr]

    def warehouse_id(self, code: str) -> UUID | None:
        if self._warehouses is None:
            self._load_warehouses()
        return self._warehouses.get(code.strip().lower())  # type: ignore[union-attr]


# ---------------------------------------------------------------------------
# Field parsers
# ---------------------------------------------------------------------------


def _normalise_header(h: str) -> str:
    return h.strip().lower().replace(" ", "_")


def _str(v: str) -> str | None:
    s = v.strip() if v else ""
    return s or None


def _parse_date(value: str, row: int, result: BulkImportResult) -> datetime | None:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value.strip(), fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    result.add_error(row, "Posting Date", f"Cannot parse date '{value}'. Use YYYY-MM-DD.")
    return None


def _parse_qty(value: str, row: int, result: BulkImportResult) -> Decimal | None:
    if not value or not value.strip():
        result.add_error(row, "Quantity", "Quantity is required.")
        return None
    try:
        d = Decimal(value.strip())
        if d <= 0:
            result.add_error(row, "Quantity", "Quantity must be greater than 0.")
            return None
        return d
    except InvalidOperation:
        result.add_error(row, "Quantity", f"'{value}' is not a valid number.")
        return None


def _parse_rate(value: str) -> Decimal | None:
    if not value or not value.strip():
        return None
    try:
        return Decimal(value.strip())
    except InvalidOperation:
        return None


# ---------------------------------------------------------------------------
# Row parser
# ---------------------------------------------------------------------------


def _parse_rows(
    rows: list[tuple[int, dict[str, str]]],
    cache: _CodeCache,
    result: BulkImportResult,
) -> list[tuple[int, StockEntryCreate]]:
    """
    Parse CSV/XLSX rows into StockEntryCreate objects.

    Rows that share the same (stock_entry_type, posting_date, posting_time,
    from_warehouse_id, to_warehouse_id, remarks) are grouped into a single
    StockEntry with multiple line items. The first row number of each group
    is used for error reporting.
    """
    # Each key maps to (first_row_num, StockEntryCreate) so we can accumulate items.
    grouped: dict[tuple, tuple[int, StockEntryCreate]] = {}

    for row_num, row in rows:
        # --- entry type ---
        entry_type = _str(row.get("stock_entry_type", ""))
        if not entry_type:
            result.add_error(row_num, "Stock Entry Type", "Stock Entry Type is required.")
            continue
        if entry_type.lower() not in VALID_ENTRY_TYPES:
            result.add_error(
                row_num,
                "Stock Entry Type",
                f"Invalid type '{entry_type}'. Valid values: {', '.join(sorted(VALID_ENTRY_TYPES))}",
            )
            continue

        # --- posting date ---
        posting_date = _parse_date(row.get("posting_date", ""), row_num, result)
        if posting_date is None:
            continue

        # --- warehouses (optional, resolved by code) ---
        from_wh_id: UUID | None = None
        to_wh_id: UUID | None = None

        from_code = _str(row.get("from_warehouse_code", ""))
        if from_code:
            from_wh_id = cache.warehouse_id(from_code)
            if from_wh_id is None:
                result.add_error(row_num, "From Warehouse Code", f"Warehouse '{from_code}' not found.")
                continue

        to_code = _str(row.get("to_warehouse_code", ""))
        if to_code:
            to_wh_id = cache.warehouse_id(to_code)
            if to_wh_id is None:
                result.add_error(row_num, "To Warehouse Code", f"Warehouse '{to_code}' not found.")
                continue

        # --- item ---
        item_code = _str(row.get("item_code", ""))
        if not item_code:
            result.add_error(row_num, "Item Code", "Item Code is required.")
            continue
        item_id = cache.item_id(item_code)
        if item_id is None:
            result.add_error(row_num, "Item Code", f"Item '{item_code}' not found.")
            continue

        # --- quantity & uom ---
        qty = _parse_qty(row.get("quantity", ""), row_num, result)
        if qty is None:
            continue

        uom = _str(row.get("uom", ""))
        if not uom:
            result.add_error(row_num, "UOM", "UOM is required.")
            continue

        # --- warehouse presence rules (fail fast, same rules as submit) ---
        entry_type_lower = entry_type.lower()
        if entry_type_lower == "material_receipt" and not to_wh_id:
            result.add_error(row_num, "To Warehouse Code", "To Warehouse Code is required for material_receipt.")
            continue
        if entry_type_lower == "material_issue" and not from_wh_id:
            result.add_error(row_num, "From Warehouse Code", "From Warehouse Code is required for material_issue.")
            continue
        if entry_type_lower in ("material_transfer", "send_to_subcontractor"):
            if not from_wh_id:
                result.add_error(row_num, "From Warehouse Code", f"From Warehouse Code is required for {entry_type_lower}.")
                continue
            if not to_wh_id:
                result.add_error(row_num, "To Warehouse Code", f"To Warehouse Code is required for {entry_type_lower}.")
                continue

        # --- optional fields ---
        basic_rate = _parse_rate(row.get("basic_rate", ""))
        valuation_rate = _parse_rate(row.get("valuation_rate", ""))
        batch_no = _str(row.get("batch_number", ""))
        description = _str(row.get("description", ""))
        remarks = _str(row.get("remarks", ""))
        posting_time = _str(row.get("posting_time", ""))

        item = StockEntryItemCreate(
            item_id=item_id,
            qty=qty,
            uom=uom,
            source_warehouse_id=from_wh_id,
            target_warehouse_id=to_wh_id,
            basic_rate=basic_rate,
            valuation_rate=valuation_rate,
            batch_no=batch_no,
            description=description,
        )

        # Group key: rows sharing the same entry-level fields become one stock entry.
        # Remarks is intentionally excluded from the key — rows with different remarks
        # still belong to the same entry; the first row's remarks value is used.
        group_key = (
            entry_type_lower,
            posting_date,
            posting_time or "",
            from_wh_id,
            to_wh_id,
        )

        if group_key in grouped:
            # Append item to the existing entry for this group.
            grouped[group_key][1].items.append(item)
        else:
            entry = StockEntryCreate(
                stock_entry_type=entry_type_lower,
                posting_date=posting_date,
                posting_time=posting_time,
                from_warehouse_id=from_wh_id,
                to_warehouse_id=to_wh_id,
                remarks=remarks,
                items=[item],
            )
            grouped[group_key] = (row_num, entry)

    return list(grouped.values())


# ---------------------------------------------------------------------------
# File parsers
# ---------------------------------------------------------------------------


def _normalise_row(raw: dict[str, str]) -> dict[str, str]:
    return {_normalise_header(k): v for k, v in raw.items()}


def _parse_csv(content: bytes) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [_normalise_header(h) for h in (reader.fieldnames or [])]
    rows = []
    for i, row in enumerate(reader, start=2):
        rows.append((i, _normalise_row(row)))
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return [], []
    raw_headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    headers = [_normalise_header(h) for h in raw_headers]
    rows = []
    for i, row in enumerate(all_rows[1:], start=2):
        values = [str(c).strip() if c is not None else "" for c in row]
        while len(values) < len(headers):
            values.append("")
        rows.append((i, dict(zip(headers, values))))
    return headers, rows


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------


class StockEntryBulkImportService:
    MAX_ROWS = 500

    def __init__(self, db: Session):
        self.db = db
        self.svc = StockEntryService(db)

    def import_from_csv(self, content: bytes, organization_id: UUID, user_id: UUID) -> BulkImportResult:
        headers, rows = _parse_csv(content)
        return self._process(headers, rows, organization_id, user_id)

    def import_from_xlsx(self, content: bytes, organization_id: UUID, user_id: UUID) -> BulkImportResult:
        headers, rows = _parse_xlsx(content)
        return self._process(headers, rows, organization_id, user_id)

    def _process(
        self,
        headers: list[str],
        rows: list[tuple[int, dict[str, str]]],
        organization_id: UUID,
        user_id: UUID,
    ) -> BulkImportResult:
        result = BulkImportResult()

        # Validate required columns present
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            # Try to give friendly names back
            friendly = {h.replace("_", " ").title() for h in missing}
            result.add_error(1, "headers", f"Missing required columns: {', '.join(sorted(friendly))}")
            return result

        # Drop fully empty rows
        data_rows = [(rn, r) for rn, r in rows if any(v.strip() for v in r.values())]
        result.total_rows = len(data_rows)

        if result.total_rows == 0:
            result.add_error(1, "file", "File contains no data rows.")
            return result

        if result.total_rows > self.MAX_ROWS:
            result.add_error(1, "file", f"Too many rows ({result.total_rows}). Maximum allowed: {self.MAX_ROWS}.")
            return result

        cache = _CodeCache(self.db, organization_id)
        entries = _parse_rows(data_rows, cache, result)

        for row_num, entry in entries:
            try:
                self.svc.create(entry, organization_id, user_id)
                result.created += 1
            except Exception as exc:
                logger.warning("Row %d: failed to create stock entry: %s", row_num, exc)
                result.add_error(row_num, "create", str(exc))

        return result


# ---------------------------------------------------------------------------
# Template generators
# ---------------------------------------------------------------------------

_SAMPLE_ROWS = [
    ["material_receipt", "2025-01-15", "09:00", "", "WH-MAIN", "ITEM-001", "Widget A opening stock", "Opening stock receipt", "100", "Pieces", "25.00", "", ""],
    ["material_receipt", "2025-01-15", "09:00", "", "WH-MAIN", "ITEM-002", "Widget B opening stock", "Opening stock receipt", "50", "Boxes", "10.00", "", "BATCH-001"],
    ["material_issue", "2025-01-16", "14:30", "WH-MAIN", "", "ITEM-001", "Issued to production", "Production run #42", "20", "Pieces", "", "", ""],
    ["material_transfer", "2025-01-17", "10:00", "WH-MAIN", "WH-SECONDARY", "ITEM-001", "", "Rebalancing stock", "30", "Pieces", "25.00", "", ""],
]


def generate_csv_template() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerows(_SAMPLE_ROWS)
    return buf.getvalue().encode("utf-8")


def generate_xlsx_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Entries"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in _SAMPLE_ROWS:
        ws.append(row)

    # Auto-width
    col_widths = [max(len(str(TEMPLATE_HEADERS[i])), max((len(str(r[i])) for r in _SAMPLE_ROWS), default=0)) + 4
                  for i in range(len(TEMPLATE_HEADERS))]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Instructions sheet
    info = wb.create_sheet("Instructions")
    info_headers = ["Column", "Required", "Description", "Valid Values / Example"]
    info.append(info_headers)
    for cell in info[1]:
        cell.font = Font(bold=True)

    instructions = [
        ("Stock Entry Type", "Yes", "Type of stock movement", "material_receipt | material_issue | material_transfer | manufacture | repack | send_to_subcontractor"),
        ("Posting Date",     "Yes", "Date of the entry",      "YYYY-MM-DD  e.g. 2025-01-15"),
        ("Posting Time",     "No",  "Time of the entry",      "HH:MM  e.g. 09:00"),
        ("From Warehouse Code", "No", "Source warehouse code (required for issue/transfer)", "e.g. WH-MAIN"),
        ("To Warehouse Code",   "No", "Target warehouse code (required for receipt/transfer)", "e.g. WH-SECONDARY"),
        ("Item Code",        "Yes", "Item code as defined in the Items master", "e.g. ITEM-001"),
        ("Description",      "No",  "Line item description",  ""),
        ("Remarks",          "No",  "Notes for the entry",    ""),
        ("Quantity",         "Yes", "Quantity (must be > 0)", "e.g. 100"),
        ("UOM",              "Yes", "Unit of measure",        "e.g. Pieces, Kg, Boxes"),
        ("Basic Rate",       "No",  "Unit cost",              "e.g. 25.00"),
        ("Valuation Rate",   "No",  "Valuation rate",         "e.g. 25.00"),
        ("Batch Number",     "No",  "Batch number if applicable", "e.g. BATCH-001"),
    ]
    for row in instructions:
        info.append(row)

    for col in info.columns:
        info.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
