"""Export service for Chart of Accounts data"""

import csv
import io
import json
import logging
from datetime import date
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from app.models.base import AccountStatus, AccountType
from app.services.report_service import ReportService

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting Chart of Accounts data in various formats.
    
    Supports:
    - CSV export
    - JSON export
    - XLSX export (Excel)
    - PDF export
    """
    
    def __init__(self, report_service: ReportService):
        """
        Initialize export service
        
        Args:
            report_service: Report service for generating data
        """
        self.report_service = report_service
    
    def export_to_csv(
        self,
        organization_id: UUID,
        account_type: Optional[AccountType] = None,
        status: Optional[AccountStatus] = None,
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """
        Export Chart of Accounts to CSV format.
        
        Args:
            organization_id: Organization UUID
            account_type: Filter by account type (optional)
            status: Filter by account status (optional)
            as_of_date: Date to calculate balances as of (defaults to today)
            
        Returns:
            CSV data as bytes
        """
        # Generate report data
        report = self.report_service.generate_chart_of_accounts_report(
            organization_id=organization_id,
            account_type=account_type,
            status=status,
            as_of_date=as_of_date
        )
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "Account Code",
            "Account Name",
            "Account Type",
            "Status",
            "Currency",
            "Is Posting Account",
            "Balance",
            "Base Currency Balance",
        ])
        
        # Write data rows
        for account in report["accounts"]:
            writer.writerow([
                account["account_code"],
                account["account_name"],
                account["account_type"],
                account["status"],
                account["currency"],
                "Yes" if account["is_posting_account"] else "No",
                f"{account['balance']:.2f}",
                f"{account['base_currency_balance']:.2f}",
            ])
        
        # Convert to bytes
        csv_data = output.getvalue().encode('utf-8')
        output.close()
        
        return csv_data
    
    def export_to_json(
        self,
        organization_id: UUID,
        account_type: Optional[AccountType] = None,
        status: Optional[AccountStatus] = None,
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """
        Export Chart of Accounts to JSON format.
        
        Args:
            organization_id: Organization UUID
            account_type: Filter by account type (optional)
            status: Filter by account status (optional)
            as_of_date: Date to calculate balances as of (defaults to today)
            
        Returns:
            JSON data as bytes
        """
        # Generate report data
        report = self.report_service.generate_chart_of_accounts_report(
            organization_id=organization_id,
            account_type=account_type,
            status=status,
            as_of_date=as_of_date
        )
        
        # Convert to JSON with pretty formatting
        json_data = json.dumps(report, indent=2, default=str)
        
        return json_data.encode('utf-8')
    
    def export_to_xlsx(
        self,
        organization_id: UUID,
        account_type: Optional[AccountType] = None,
        status: Optional[AccountStatus] = None,
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """
        Export Chart of Accounts to XLSX format (Excel).
        
        Args:
            organization_id: Organization UUID
            account_type: Filter by account type (optional)
            status: Filter by account status (optional)
            as_of_date: Date to calculate balances as of (defaults to today)
            
        Returns:
            XLSX data as bytes
        """
        # Generate report data
        report = self.report_service.generate_chart_of_accounts_report(
            organization_id=organization_id,
            account_type=account_type,
            status=status,
            as_of_date=as_of_date
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Chart of Accounts"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            "Account Code",
            "Account Name",
            "Account Type",
            "Status",
            "Currency",
            "Is Posting Account",
            "Balance",
            "Base Currency Balance",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_num, account in enumerate(report["accounts"], 2):
            ws.cell(row=row_num, column=1, value=account["account_code"])
            ws.cell(row=row_num, column=2, value=account["account_name"])
            ws.cell(row=row_num, column=3, value=account["account_type"])
            ws.cell(row=row_num, column=4, value=account["status"])
            ws.cell(row=row_num, column=5, value=account["currency"])
            ws.cell(row=row_num, column=6, value="Yes" if account["is_posting_account"] else "No")
            ws.cell(row=row_num, column=7, value=account["balance"])
            ws.cell(row=row_num, column=8, value=account["base_currency_balance"])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        xlsx_data = output.getvalue()
        output.close()
        
        return xlsx_data
    
    def export_to_pdf(
        self,
        organization_id: UUID,
        account_type: Optional[AccountType] = None,
        status: Optional[AccountStatus] = None,
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """
        Export Chart of Accounts to PDF format.
        
        Args:
            organization_id: Organization UUID
            account_type: Filter by account type (optional)
            status: Filter by account status (optional)
            as_of_date: Date to calculate balances as of (defaults to today)
            
        Returns:
            PDF data as bytes
        """
        # Generate report data
        report = self.report_service.generate_chart_of_accounts_report(
            organization_id=organization_id,
            account_type=account_type,
            status=status,
            as_of_date=as_of_date
        )
        
        # Create PDF in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        # Container for PDF elements
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )
        
        # Add title
        title = Paragraph("Chart of Accounts Report", title_style)
        elements.append(title)
        
        # Add report metadata
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
        )
        
        metadata_text = f"As of: {report['as_of_date']}"
        if report['filters']['account_type']:
            metadata_text += f" | Type: {report['filters']['account_type']}"
        if report['filters']['status']:
            metadata_text += f" | Status: {report['filters']['status']}"
        
        metadata = Paragraph(metadata_text, metadata_style)
        elements.append(metadata)
        elements.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = [
            ["Code", "Name", "Type", "Status", "Currency", "Posting", "Balance"]
        ]
        
        for account in report["accounts"]:
            table_data.append([
                account["account_code"],
                account["account_name"][:30] + "..." if len(account["account_name"]) > 30 else account["account_name"],
                account["account_type"][:10] if account["account_type"] else "",
                account["status"][:10] if account["status"] else "",
                account["currency"],
                "Yes" if account["is_posting_account"] else "No",
                f"{account['balance']:.2f}",
            ])
        
        # Create table
        table = Table(table_data, colWidths=[
            0.8*inch,  # Code
            2.2*inch,  # Name
            0.8*inch,  # Type
            0.7*inch,  # Status
            0.6*inch,  # Currency
            0.6*inch,  # Posting
            0.9*inch,  # Balance
        ])
        
        # Style the table
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Balance column right-aligned
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add footer with total count
        elements.append(Spacer(1, 20))
        footer_text = f"Total Accounts: {report['total_accounts']}"
        footer = Paragraph(footer_text, metadata_style)
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        pdf_data = output.getvalue()
        output.close()
        
        return pdf_data
    
    def export_trial_balance_to_pdf(
        self,
        organization_id: UUID,
        account_type: Optional[AccountType] = None,
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """
        Export Trial Balance to PDF format.
        
        Args:
            organization_id: Organization UUID
            account_type: Filter by account type (optional)
            as_of_date: Date to calculate balances as of (defaults to today)
            
        Returns:
            PDF data as bytes
        """
        # Generate trial balance report
        report = self.report_service.generate_trial_balance(
            organization_id=organization_id,
            account_type=account_type,
            as_of_date=as_of_date
        )
        
        # Create PDF in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        # Container for PDF elements
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )
        
        # Add title
        title = Paragraph("Trial Balance Report", title_style)
        elements.append(title)
        
        # Add report metadata
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER,
        )
        
        metadata_text = f"As of: {report['as_of_date']}"
        if report['filters']['account_type']:
            metadata_text += f" | Type: {report['filters']['account_type']}"
        
        metadata = Paragraph(metadata_text, metadata_style)
        elements.append(metadata)
        elements.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = [
            ["Code", "Name", "Type", "Debit", "Credit"]
        ]
        
        for account in report["accounts"]:
            table_data.append([
                account["account_code"],
                account["account_name"][:35] + "..." if len(account["account_name"]) > 35 else account["account_name"],
                account["account_type"][:10] if account["account_type"] else "",
                f"{account['debit_balance']:.2f}" if account['debit_balance'] > 0 else "",
                f"{account['credit_balance']:.2f}" if account['credit_balance'] > 0 else "",
            ])
        
        # Add totals row
        table_data.append([
            "",
            "TOTAL",
            "",
            f"{report['total_debits']:.2f}",
            f"{report['total_credits']:.2f}",
        ])
        
        # Create table
        table = Table(table_data, colWidths=[
            1.0*inch,  # Code
            2.8*inch,  # Name
            1.0*inch,  # Type
            1.0*inch,  # Debit
            1.0*inch,  # Credit
        ])
        
        # Style the table
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (4, -1), 'RIGHT'),  # Debit/Credit columns right-aligned
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Totals row
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add balance status
        elements.append(Spacer(1, 20))
        balance_status = "✓ Trial Balance is BALANCED" if report['is_balanced'] else f"✗ Trial Balance is OUT OF BALANCE (Difference: {report['difference']:.2f})"
        status_color = colors.green if report['is_balanced'] else colors.red
        
        status_style = ParagraphStyle(
            'Status',
            parent=styles['Normal'],
            fontSize=11,
            textColor=status_color,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
        )
        
        status = Paragraph(balance_status, status_style)
        elements.append(status)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        pdf_data = output.getvalue()
        output.close()
        
        return pdf_data
