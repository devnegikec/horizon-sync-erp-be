"""Service layer for QSeal module"""

import logging
import uuid
from datetime import UTC, datetime
from uuid import UUID

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.repositories.qseal_repository import QSealRepository
from app.schemas.qseal import (
    QSealChildCreate,
    QSealMapRequest,
    QSealParentCreate,
    QSealScanRequest,
)

logger = logging.getLogger(__name__)


class QSealService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = QSealRepository(db)

    def _to_response_dict(self, node) -> dict:
        return {
            "id": node.id,
            "organization_id": node.organization_id,
            "qseal_type": node.qseal_type,
            "name": node.name,
            "capacity": node.capacity,
            "serial_number": node.serial_number,
            "qseal_code_link": node.qseal_code_link,
            "app_cascade_map": node.app_cascade_map,
            "parent_id": node.parent_id,
            "parent_app_id": node.parent_app_id,
            "children_count": self.repo.count_children(node.id),
            "created_at": node.created_at,
        }

    def _paginate(self, items, total, page, page_size) -> dict:
        total_pages = max(1, (total + page_size - 1) // page_size)
        return {
            "page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
        }

    # ── Parent QSeal ──────────────────────────────────────────────────────────

    def create_parent(self, data: QSealParentCreate, organization_id: UUID):
        serial = self.repo.generate_serial(prefix="QSL")
        payload = data.model_dump(exclude={"extra_data"})
        payload["organization_id"] = organization_id
        payload["serial_number"] = serial
        payload["qseal_code_link"] = f"/qseal/{serial}"
        node = self.repo.create_node(payload)
        logger.info(
            "[QSEAL] parent created id=%s serial=%s org=%s",
            node.id,
            serial,
            organization_id,
        )
        return self._to_response_dict(node)

    def list_parents(
        self,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 20,
        qseal_type: str | None = None,
    ):
        items, total = self.repo.list_roots(
            organization_id, page, page_size, qseal_type
        )
        return {
            "nodes": [self._to_response_dict(n) for n in items],
            "pagination": self._paginate(items, total, page, page_size),
        }

    def get_parent(self, node_id: UUID, organization_id: UUID):
        node = self.repo.get_by_id(node_id, organization_id)
        if not node:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="QSeal node not found"
            )
        return self._to_response_dict(node)

    # ── Child QSeal ───────────────────────────────────────────────────────────

    def create_child(
        self, parent_id: UUID, data: QSealChildCreate, organization_id: UUID
    ):
        parent = self.repo.get_by_id(parent_id, organization_id)
        if not parent:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Parent QSeal node not found",
            )

        # Capacity check
        current_children = self.repo.count_children(parent_id)
        if parent.capacity and current_children >= parent.capacity:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Parent node is at full capacity ({parent.capacity})",
            )

        serial = self.repo.generate_serial(prefix="QSL")
        payload = data.model_dump(exclude={"extra_data"})
        payload["organization_id"] = organization_id
        payload["parent_id"] = parent_id
        payload["serial_number"] = serial
        payload["qseal_code_link"] = f"/qseal/{serial}"
        node = self.repo.create_node(payload)
        logger.info(
            "[QSEAL] child created id=%s parent=%s org=%s",
            node.id,
            parent_id,
            organization_id,
        )
        return self._to_response_dict(node)

    def list_children(
        self,
        parent_id: UUID,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 50,
    ):
        # Validate parent exists
        parent = self.repo.get_by_id(parent_id, organization_id)
        if not parent:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Parent QSeal node not found",
            )
        items, total = self.repo.list_children(
            parent_id, organization_id, page, page_size
        )
        return {
            "children": [self._to_response_dict(n) for n in items],
            "pagination": self._paginate(items, total, page, page_size),
        }

    # ── Map QSeals ────────────────────────────────────────────────────────────

    def map_children(
        self, parent_id: UUID, req: QSealMapRequest, organization_id: UUID
    ):
        parent = self.repo.get_by_id(parent_id, organization_id)
        if not parent:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Parent QSeal node not found",
            )

        current_children = self.repo.count_children(parent_id)
        if (
            parent.capacity
            and (current_children + len(req.child_ids)) > parent.capacity
        ):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"Mapping would exceed parent capacity ({parent.capacity}). "
                    f"Current: {current_children}, Requested: {len(req.child_ids)}"
                ),
            )

        mapped = self.repo.map_children(parent_id, req.child_ids, organization_id)
        logger.info(
            "[QSEAL] map_children parent=%s requested=%d mapped=%d ids=%s",
            parent_id,
            len(req.child_ids),
            mapped,
            req.child_ids,
        )
        return {
            "parent_id": parent_id,
            "mapped_count": mapped,
            "message": f"Successfully mapped {mapped} child QSeal(s) to parent.",
        }

    # ── QSeal Scan ────────────────────────────────────────────────────────────

    def record_scan(self, req: QSealScanRequest, organization_id: UUID):
        # Resolve the QSealTrack parent strictly within the supplied tenant.
        # The /scan endpoint is public and organization_id is caller-supplied,
        # so a global fallback here would let a caller resolve (and expose)
        # another tenant's QSeal node.
        node = self.repo.get_by_serial(req.serial_number, organization_id)
        is_parent = True

        # 2. Fallback: try QSealParameters (child units from ProductItems)
        if not node:
            from app.models.qseal import QSealParameters

            child = (
                self.db.query(QSealParameters)
                .filter(
                    QSealParameters.serial_number == req.serial_number,
                    QSealParameters.organization_id == organization_id,
                )
                .first()
            )
            if child:
                is_parent = False
                # Build a pseudo-node response with parent info
                parent_node = None
                parent_serial = None
                if child.parent_id:
                    parent_node = self.repo.get_by_id(child.parent_id, organization_id)
                    parent_serial = parent_node.serial_number if parent_node else None

                # Record scan event
                scan_payload = {
                    "organization_id": organization_id,
                    "serial_number": req.serial_number,
                    "scan_timestamp": datetime.now(UTC),
                    "device_type": req.device_type,
                    "os": req.os,
                    "browser": req.browser,
                    "ip_address": req.ip_address,
                    "latitude": req.latitude,
                    "longitude": req.longitude,
                    "city": req.city,
                    "state": req.state,
                    "country": req.country,
                    "extra_data": req.extra_data,
                }
                self.repo.record_scan(scan_payload)

                logger.info(
                    "[QSEAL] child scan recorded serial=%s org=%s parent=%s",
                    req.serial_number,
                    organization_id,
                    parent_serial,
                )
                return {
                    "node_id": child.id,
                    "serial_number": child.serial_number,
                    "qseal_type": "child_unit",
                    "name": f"Unit {child.serial_number or ''}",
                    "parent_id": child.parent_id,
                    "parent_serial": parent_serial,
                    "children_count": 0,
                    "message": f"Child QSeal unit scanned. Parent: {parent_serial or 'none'}.",
                }

        # 3. Not found in either table
        if not node:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=(
                    f"No QSeal node found for serial '{req.serial_number}' "
                    f"(organization_id={organization_id})"
                ),
            )

        # Record scan for parent node
        scan_payload = {
            "organization_id": organization_id,
            "serial_number": req.serial_number,
            "scan_timestamp": datetime.now(UTC),
            "device_type": req.device_type,
            "os": req.os,
            "browser": req.browser,
            "ip_address": req.ip_address,
            "latitude": req.latitude,
            "longitude": req.longitude,
            "city": req.city,
            "state": req.state,
            "country": req.country,
            "extra_data": req.extra_data,
        }
        self.repo.record_scan(scan_payload)

        children_count = self.repo.count_children(node.id)
        logger.info(
            "[QSEAL] scan recorded serial=%s node=%s org=%s",
            req.serial_number,
            node.id,
            organization_id,
        )

        return {
            "node_id": node.id,
            "serial_number": node.serial_number,
            "qseal_type": node.qseal_type,
            "name": node.name,
            "parent_id": node.parent_id,
            "children_count": children_count,
            "message": f"QSeal scan recorded for {node.qseal_type or 'node'} '{node.name}'.",
        }

    # ── QSeal History ─────────────────────────────────────────────────────────

    def get_scan_history(
        self,
        organization_id: UUID,
        serial_number: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ):
        items, total = self.repo.list_scan_history(
            organization_id, serial_number, page, page_size
        )
        return {
            "events": items,
            "pagination": self._paginate(items, total, page, page_size),
        }

    # ── Label Download ────────────────────────────────────────────────────────

    def get_labels(self, parent_id: UUID, organization_id: UUID):
        """Return all child nodes under a parent as label data for printing."""
        parent = self.repo.get_by_id(parent_id, organization_id)
        if not parent:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Parent QSeal node not found",
            )

        # Fetch all children (no pagination — label batch)
        items, total = self.repo.list_children(
            parent_id, organization_id, page=1, page_size=10000
        )
        labels = [
            {
                "id": str(child.id),
                "serial_number": child.serial_number,
                "qseal_type": child.qseal_type,
                "name": child.name,
                "qseal_code_link": child.qseal_code_link,
                "parent_serial": parent.serial_number,
            }
            for child in items
        ]
        return {
            "parent_id": parent_id,
            "labels": labels,
            "total": total,
        }

    # ── Parent with Linked Units (for inbound/receiving) ──────────────────────

    def get_parent_with_linked_units(
        self, parent_id: UUID, organization_id: UUID
    ) -> dict:
        """Return a parent QSeal node with all its linked QSealParameters children.

        Used by mobile app for inbound: scan parent QR → see all linked units
        → create receiving slip.
        """
        from app.models.product_item import ProductItem
        from app.models.qr_product import QRProduct
        from app.models.qseal import QSealParameters

        parent = self.repo.get_by_id(parent_id, organization_id)
        if not parent:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Parent QSeal node not found",
            )

        # The parent is already org-scoped, so linked units are resolved within
        # the same organization.
        parent_org = parent.organization_id or organization_id

        # Fetch linked QSealParameters with ProductItem + QRProduct joins
        linked = (
            self.db.query(QSealParameters, ProductItem, QRProduct)
            .outerjoin(
                ProductItem,
                (ProductItem.serial_number == QSealParameters.serial_number)
                & (ProductItem.organization_id == parent_org),
            )
            .outerjoin(
                QRProduct,
                QRProduct.id == QSealParameters.product_id,
            )
            .filter(
                QSealParameters.parent_id == parent_id,
                QSealParameters.organization_id == parent_org,
            )
            .order_by(QSealParameters.created_at.asc())
            .all()
        )

        units = []
        for param, item, product in linked:
            units.append(
                {
                    "id": param.id,
                    "serial_number": param.serial_number,
                    "product_name": product.name if product else None,
                    "product_sku": product.sku if product else None,
                    "manufacturing_date": str(param.manufacturing_date)
                    if param.manufacturing_date
                    else None,
                    "expiry_date": str(param.expiry_date)
                    if param.expiry_date
                    else None,
                    "manufacturing_unit": param.manufacturing_unit,
                    "dispatch_batch": param.dispatch_batch,
                    "destination_market": param.destination_market,
                    "mrp": float(param.mrp) if param.mrp else None,
                    "currency": param.currency,
                    "batch_size": param.batch_size,
                    "qseal_cascade": param.qseal_cascade or False,
                    "product_item_url": item.token_id if item else None,
                    "product_item_scan_count": item.scan_count if item else 0,
                    "extra_data": param.extra_data,
                }
            )

        result = self._to_response_dict(parent)
        result["linked_units"] = units
        logger.info(
            "[QSEAL] parent detail with linked units parent=%s units=%d",
            parent_id,
            len(units),
        )
        return result

    # ── Block-based Parent QSeal ──────────────────────────────────────────────

    def get_parents_by_block(
        self,
        block_id: UUID,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 20,
    ):
        """List all QSealTrack parent nodes created for a QR block's master packs."""
        from app.models.qr_block import QRBlock
        from app.models.qseal import QSealTrack

        # Validate block exists and belongs to org
        block = (
            self.db.query(QRBlock)
            .filter(QRBlock.id == block_id, QRBlock.organization_id == organization_id)
            .first()
        )
        if not block:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="QR block not found",
            )

        if not block.master_pack_enabled:
            return {
                "nodes": [],
                "pagination": self._paginate([], 0, page, page_size),
                "message": "Master pack is not enabled for this block.",
            }

        # Find all distinct parent QSealTracks via QSealParameters
        from app.models.qseal import QSealParameters

        parent_ids_subq = (
            self.db.query(QSealParameters.parent_id)
            .filter(
                QSealParameters.block_id == block_id,
                QSealParameters.organization_id == organization_id,
                QSealParameters.parent_id.isnot(None),
            )
            .distinct()
            .subquery()
        )

        q = self.db.query(QSealTrack).filter(
            QSealTrack.id.in_(self.db.query(parent_ids_subq.c.parent_id))
        )

        total = q.count()
        items = (
            q.order_by(QSealTrack.created_at.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return {
            "nodes": [self._to_response_dict(n) for n in items],
            "pagination": self._paginate(items, total, page, page_size),
            "block_batch": block.batch,
        }

    def get_parents_excel(
        self, block_id: UUID, organization_id: UUID
    ) -> tuple[bytes, str]:
        """Generate an Excel file with parent QSeal QR codes for a block.

        Embeds QR code images for mobile app scanning only when the block's
        ``qr_image`` flag is enabled. Returns (excel_bytes, filename).
        """
        from io import BytesIO

        import qrcode
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        from app.config import settings
        from app.models.qr_block import QRBlock
        from app.models.qseal import QSealParameters, QSealTrack

        def _embed_qr(ws, url: str, row: int, col: int, size: int = 150) -> None:
            if not url:
                return
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=2,
            )
            qr.add_data(url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            buf = BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            xl_img = XLImage(buf)
            xl_img.width = size
            xl_img.height = size
            cell_ref = f"{get_column_letter(col)}{row}"
            ws.add_image(xl_img, cell_ref)

        # Validate block
        block = (
            self.db.query(QRBlock)
            .filter(QRBlock.id == block_id, QRBlock.organization_id == organization_id)
            .first()
        )
        if not block:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="QR block not found",
            )

        if not block.master_pack_enabled:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Master pack is not enabled for this block.",
            )

        # Get parent nodes via parameters
        parent_ids = (
            self.db.query(QSealParameters.parent_id)
            .filter(
                QSealParameters.block_id == block_id,
                QSealParameters.organization_id == organization_id,
                QSealParameters.parent_id.isnot(None),
            )
            .distinct()
            .all()
        )
        parent_id_list = [p[0] for p in parent_ids]

        if not parent_id_list:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No parent QSeal nodes found for this block.",
            )

        parents = (
            self.db.query(QSealTrack)
            .filter(QSealTrack.id.in_(parent_id_list))
            .order_by(QSealTrack.created_at.asc())
            .all()
        )

        # Build Excel — embed QR code images only when the block requested them
        include_images = bool(block.qr_image)
        wb = Workbook()
        ws = wb.active
        ws.title = "QSeal Parent QR Codes"

        # Headers: QR URL, [QR Code image], Serial, Name, Capacity
        headers = (
            ["QR URL", "QR Code", "Serial Number", "Name", "Capacity"]
            if include_images
            else ["QR URL", "Serial Number", "Name", "Capacity"]
        )
        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions[get_column_letter(1)].width = 55  # QR URL
        if include_images:
            ws.column_dimensions[get_column_letter(2)].width = 24  # QR Code image
        serial_col = 3 if include_images else 2
        name_col = 4 if include_images else 3
        capacity_col = 5 if include_images else 4
        ws.column_dimensions[get_column_letter(serial_col)].width = 18  # Serial
        ws.column_dimensions[get_column_letter(name_col)].width = 25  # Name
        ws.column_dimensions[get_column_letter(capacity_col)].width = 15  # Capacity

        qr_size = 150
        base_url = settings.qr_base_url or f"https://{settings.qr_domain}"

        for row_idx, parent in enumerate(parents, 2):
            serial = parent.serial_number or ""
            qr_url = f"{base_url}/qseal/{serial}" if serial else ""

            if include_images:
                # Row height for QR image
                ws.row_dimensions[row_idx].height = 115

            ws.cell(row=row_idx, column=1, value=qr_url)  # QR URL
            if include_images:
                _embed_qr(ws, qr_url, row_idx, 2, qr_size)  # QR Code image
            ws.cell(row=row_idx, column=serial_col, value=serial)  # Serial Number
            ws.cell(row=row_idx, column=name_col, value=parent.name or "")
            ws.cell(row=row_idx, column=capacity_col, value=parent.capacity or 0)

        # Save
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"qseal_parents_{block.batch}.xlsx"
        logger.info(
            "[QSEAL] parent excel generated block=%s parents=%d images=%s",
            block_id,
            len(parents),
            include_images,
        )
        return buf.getvalue(), filename

    # ── Auto-link (automatic cascade / aggregation) ─────────────────────────

    def auto_link_block(
        self,
        block_id: UUID,
        organization_id: UUID,
        user_id: UUID | None = None,
        master_pack_size: int | None = None,
    ) -> dict:
        """Automatically cascade a completed block's items into master packs.

        Groups the block's generated ProductItems into chunks of
        ``master_pack_size``, creating a QSealTrack (shipper) parent per chunk
        and linking each item via a QSealParameters row. Existing linkage for
        the block is removed first so the operation is idempotent (re-cascade).
        """
        from app.models.product_item import ProductItem
        from app.models.qr_block import QRBlock
        from app.models.qseal import QSealParameters, QSealTrack

        block = (
            self.db.query(QRBlock)
            .filter(
                QRBlock.id == block_id,
                QRBlock.organization_id == organization_id,
            )
            .with_for_update()
            .first()
        )
        if not block:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="QR block not found",
            )

        if block.status != "completed":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Block is not ready (status: {block.status})",
            )

        pack_size = master_pack_size or block.master_pack_size
        if not pack_size or pack_size <= 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="master_pack_size is required to auto-link this block",
            )

        items = (
            self.db.query(ProductItem)
            .filter(
                ProductItem.block_id == block_id,
                ProductItem.organization_id == organization_id,
                ProductItem.deleted_at.is_(None),
            )
            .order_by(
                ProductItem.created_at.asc(),
                ProductItem.serial_number.asc(),
            )
            .all()
        )
        if not items:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Block has no generated items to link",
            )

        # Remove previous linkage for this block so re-cascading is idempotent.
        existing_params = (
            self.db.query(QSealParameters)
            .filter(
                QSealParameters.block_id == block_id,
                QSealParameters.organization_id == organization_id,
            )
            .all()
        )
        orphan_parent_ids = {p.parent_id for p in existing_params if p.parent_id}
        for p in existing_params:
            self.db.delete(p)
        self.db.flush()

        # Delete orphaned QSealTrack parents no longer referenced by any child.
        for parent_id in orphan_parent_ids:
            still_referenced = (
                self.db.query(QSealParameters.id)
                .filter(QSealParameters.parent_id == parent_id)
                .first()
            )
            has_track_children = (
                self.db.query(QSealTrack.id)
                .filter(QSealTrack.parent_id == parent_id)
                .first()
            )
            if not still_referenced and not has_track_children:
                orphan = (
                    self.db.query(QSealTrack).filter(QSealTrack.id == parent_id).first()
                )
                if orphan:
                    self.db.delete(orphan)

        now = datetime.now(UTC)
        parent_count = 0
        for chunk_start in range(0, len(items), pack_size):
            chunk = items[chunk_start : chunk_start + pack_size]
            if not chunk:
                continue

            parent_serial = self.repo.generate_serial(prefix="QSL")
            parent_name = f"MP-{block.batch[:10]}-{parent_count + 1}"[:20]
            parent_node = QSealTrack(
                id=uuid.uuid4(),
                organization_id=organization_id,
                qseal_type="shipper",
                name=parent_name,
                capacity=pack_size,
                serial_number=parent_serial,
                qseal_code_link=f"/qseal/{parent_serial}",
                app_cascade_map=False,
                created_at=now,
            )
            self.db.add(parent_node)
            self.db.flush()

            for item in chunk:
                self.db.add(
                    QSealParameters(
                        id=uuid.uuid4(),
                        organization_id=organization_id,
                        product_id=block.product_id,
                        block_id=block.id,
                        serial_number=item.serial_number,
                        manufacturing_date=block.manufacture_date or now.date(),
                        expiry_date=block.expiry_date or now.date(),
                        manufacturing_unit="",
                        dispatch_batch=block.batch,
                        batch_size=pack_size,
                        qseal_settings=False,
                        qseal_cascade=False,
                        parent_id=parent_node.id,
                        extra_data={
                            "item_id": str(item.id),
                            "master_pack_index": parent_count + 1,
                        },
                        created_by=user_id,
                        created_at=now,
                    )
                )

            parent_count += 1
            logger.info(
                "[QSEAL] auto-link parent created serial=%s block=%s org=%s items=%d",
                parent_serial,
                block.id,
                organization_id,
                len(chunk),
            )

        block.master_pack_enabled = True
        block.master_pack_size = pack_size
        block.extra_data = (block.extra_data or {}) | {
            "qseal_parent_count": parent_count
        }
        self.db.commit()

        logger.info(
            "[QSEAL] auto-link complete block=%s parents=%d items=%d",
            block.id,
            parent_count,
            len(items),
        )
        return {
            "block_id": block.id,
            "batch": block.batch,
            "master_pack_size": pack_size,
            "parent_count": parent_count,
            "linked_item_count": len(items),
            "message": (
                f"Auto-linked {len(items)} items into "
                f"{parent_count} master pack parent(s)."
            ),
        }

    # ── Aggregation log ─────────────────────────────────────────────────────

    def list_aggregation(
        self,
        organization_id: UUID,
        block_id: UUID | None = None,
        page: int = 1,
        page_size: int = 50,
        grouped: bool = False,
    ) -> dict:
        """List the aggregation (cascading) log.

        Flat mode (default) returns one row per child unit. Grouped mode nests
        each child under its parent (master-pack) box so the parent-child link
        is visible at a glance; unlinked units are returned separately.
        """
        from sqlalchemy import func

        from app.models.product_item import ProductItem
        from app.models.qr_block import QRBlock
        from app.models.qseal import QSealParameters, QSealTrack

        q = (
            self.db.query(ProductItem, QSealParameters, QSealTrack, QRBlock)
            .outerjoin(
                QSealParameters,
                (QSealParameters.serial_number == ProductItem.serial_number)
                & (QSealParameters.block_id == ProductItem.block_id),
            )
            .outerjoin(QSealTrack, QSealTrack.id == QSealParameters.parent_id)
            .outerjoin(QRBlock, QRBlock.id == ProductItem.block_id)
            .filter(
                ProductItem.organization_id == organization_id,
                ProductItem.deleted_at.is_(None),
            )
        )
        if block_id:
            q = q.filter(ProductItem.block_id == block_id)

        total = q.count()

        if grouped:
            rows = q.order_by(
                QSealTrack.serial_number.asc(),
                ProductItem.created_at.asc(),
                ProductItem.serial_number.asc(),
            ).all()
            return self._build_grouped_aggregation(rows, page, page_size)

        rows = (
            q.order_by(
                ProductItem.created_at.asc(),
                ProductItem.serial_number.asc(),
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        # Per-parent linked child count — used to spot over/under aggregation.
        count_rows = (
            self.db.query(
                QSealParameters.parent_id,
                func.count(QSealParameters.id),
            )
            .filter(
                QSealParameters.organization_id == organization_id,
                QSealParameters.parent_id.isnot(None),
            )
            .group_by(QSealParameters.parent_id)
            .all()
        )
        parent_counts = {parent_id: count for parent_id, count in count_rows}

        items = []
        for item, qsp, parent, blk in rows:
            linked = bool(qsp and qsp.parent_id and parent is not None)
            items.append(
                {
                    "id": item.id,
                    "block_id": item.block_id,
                    "batch": blk.batch if blk else None,
                    "child_serial": item.serial_number,
                    "activated": bool(item.qr_active),
                    "scan_count": item.scan_count or 0,
                    "linked": linked,
                    "parent_id": parent.id if linked else None,
                    "parent_serial": parent.serial_number if parent else None,
                    "parent_name": parent.name if parent else None,
                    "parent_type": parent.qseal_type if parent else None,
                    "parent_capacity": parent.capacity if parent else None,
                    "parent_linked_count": (
                        parent_counts.get(parent.id, 0) if parent else None
                    ),
                    "created_at": item.created_at,
                }
            )

        return {
            "items": items,
            "pagination": self._paginate(rows, total, page, page_size),
        }

    def _build_grouped_aggregation(self, rows, page=1, page_size=50) -> dict:
        """Group aggregation rows by parent (master-pack) box.

        Children are nested under their parent; units without a parent are
        returned in ``unlinked``. Pagination applies to the parent groups.
        """
        groups: dict[UUID, dict] = {}
        unlinked: list[dict] = []

        for item, qsp, parent, blk in rows:
            linked = bool(qsp and qsp.parent_id and parent is not None)
            if linked:
                group = groups.get(parent.id)
                if group is None:
                    group = {
                        "parent_id": parent.id,
                        "parent_serial": parent.serial_number,
                        "parent_name": parent.name,
                        "parent_type": parent.qseal_type,
                        "parent_capacity": parent.capacity,
                        "children": [],
                    }
                    groups[parent.id] = group
                group["children"].append(
                    {
                        "id": item.id,
                        "block_id": item.block_id,
                        "batch": blk.batch if blk else None,
                        "child_serial": item.serial_number,
                        "activated": bool(item.qr_active),
                        "scan_count": item.scan_count or 0,
                        "created_at": item.created_at,
                    }
                )
            else:
                unlinked.append(
                    {
                        "id": item.id,
                        "block_id": item.block_id,
                        "batch": blk.batch if blk else None,
                        "child_serial": item.serial_number,
                        "activated": bool(item.qr_active),
                        "scan_count": item.scan_count or 0,
                        "linked": False,
                        "parent_id": None,
                        "parent_serial": None,
                        "parent_name": None,
                        "parent_type": None,
                        "parent_capacity": None,
                        "parent_linked_count": None,
                        "created_at": item.created_at,
                    }
                )

        groups_list = []
        for group in groups.values():
            group["linked_count"] = len(group["children"])
            groups_list.append(group)

        total_groups = len(groups_list)
        start = (page - 1) * page_size
        page_groups = groups_list[start : start + page_size]

        return {
            "groups": page_groups,
            "unlinked": unlinked,
            "pagination": self._paginate(page_groups, total_groups, page, page_size),
        }
