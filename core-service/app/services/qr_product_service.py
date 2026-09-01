"""Service layer for QR Products module"""

import hashlib
import logging
import secrets
import string
import uuid
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from uuid import UUID

from fastapi import HTTPException, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.config import settings
from app.models.product_item import ProductItem
from app.models.qr_block import QRBlock
from app.models.qr_product import QRProduct
from app.repositories.qr_product_repository import (
    ProductItemRepository,
    QRBlockRepository,
    QRProductRepository,
)
from app.repositories.qr_product_setting_repository import QRProductSettingRepository
from app.repositories.sku_repository import ProductSKURepository
from app.schemas.qr_product import (
    AuthenticateRequest,
    QRActivationParamsCreate,
    QRBlockCreate,
    QRProductCreate,
    QRProductUpdate,
    QRType,
    QRValidateRequest,
    SerialNumberType,
    normalize_qr_type,
    normalize_serial_number_type,
)
from app.services.credit_service import CreditService
from app.services.key_service import KeyService
from app.services.qr_shortener import QRShortener
from app.utils.serial_generators import (
    build_qr_url,
    generate_r4dan,
    generate_r6dan,
    generate_r8dan,
    sequential_s8dn,
    sequential_s10dn,
    sign_qr_item,
)

logger = logging.getLogger(__name__)


def _build_excel(  # noqa: C901
    rows: list[dict],
    qr_type: str,
    include_qr_images: bool = False,
) -> bytes:
    """Build a QR workbook, optionally embedding PNG QR images."""
    from io import BytesIO

    import qrcode
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as WorksheetImage
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "QR Codes"

    normalized_type = normalize_qr_type(qr_type) or QRType.DYNAMIC
    if normalized_type == QRType.DUAL:
        headers = ["URL (Overt)", "URL (Covert)", "Serial Number"]
        image_fields = [("QR (Overt)", "overt_url"), ("QR (Covert)", "covert_url")]
    elif normalized_type == QRType.SECURE_CODE:
        headers = ["QR URL", "Serial Number", "Secret Code"]
        image_fields = [("QR Image", "primary_url")]
    else:
        headers = ["QR URL", "Serial Number"]
        image_fields = [("QR Image", "primary_url")]

    if include_qr_images:
        headers.extend(label for label, _field in image_fields)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    image_streams: list[BytesIO] = []
    for row_idx, item in enumerate(rows, 2):
        if normalized_type == QRType.DUAL:
            ws.cell(row=row_idx, column=1, value=item.get("overt_url", ""))
            ws.cell(row=row_idx, column=2, value=item.get("covert_url", ""))
            ws.cell(row=row_idx, column=3, value=item.get("serial", ""))
        elif normalized_type == QRType.SECURE_CODE:
            ws.cell(row=row_idx, column=1, value=item.get("primary_url", ""))
            ws.cell(row=row_idx, column=2, value=item.get("serial", ""))
            ws.cell(row=row_idx, column=3, value=item.get("secret_code", ""))
        else:
            ws.cell(row=row_idx, column=1, value=item.get("primary_url", ""))
            ws.cell(row=row_idx, column=2, value=item.get("serial", ""))

        for url_column in range(1, 3 if normalized_type == QRType.DUAL else 2):
            cell = ws.cell(row=row_idx, column=url_column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

        if include_qr_images:
            first_image_column = len(headers) - len(image_fields) + 1
            for offset, (_label, field) in enumerate(image_fields):
                url = item.get(field, "")
                if not url:
                    continue
                stream = BytesIO()
                qrcode.make(url).save(stream, format="PNG")
                stream.seek(0)
                image_streams.append(stream)
                image = WorksheetImage(stream)
                image.width = 96
                image.height = 96
                anchor = ws.cell(
                    row=row_idx,
                    column=first_image_column + offset,
                ).coordinate
                ws.add_image(image, anchor)
            ws.row_dimensions[row_idx].height = 75

    for column in range(1, len(headers) + 1):
        is_image_column = (
            "QR" in headers[column - 1] and "URL" not in headers[column - 1]
        )
        ws.column_dimensions[ws.cell(row=1, column=column).column_letter].width = (
            18 if is_image_column else 42
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class QRProductService:
    def __init__(self, db: Session):
        self.db = db
        self.product_repo = QRProductRepository(db)
        self.product_setting_repo = QRProductSettingRepository(db)
        self.sku_repo = ProductSKURepository(db)
        self.block_repo = QRBlockRepository(db)
        self.item_repo = ProductItemRepository(db)
        self.credit_service = CreditService(db)
        self.qr_shortener = QRShortener()
        self.key_service = (
            KeyService(settings.brand_key_encryption_secret)
            if settings.brand_key_encryption_secret
            else None
        )

    # ── Products ──────────────────────────────────────────────────────────────

    def _validate_shelf_life_setting(
        self,
        setting_id: UUID | None,
        organization_id: UUID,
        *,
        allow_inactive: bool = False,
    ) -> list[dict]:
        if setting_id is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Shelf life setting is required",
            )

        setting = self.product_setting_repo.get_by_id(setting_id, organization_id)
        if not setting:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Shelf life setting not found",
            )
        if setting.setting_type != "shelf_life":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Selected setting is not a shelf life setting",
            )
        if not setting.is_active and not allow_inactive:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Selected shelf life setting is inactive",
            )

    def _validate_serial_prefix_setting(
        self,
        setting_id: UUID | None,
        organization_id: UUID,
        *,
        allow_inactive: bool = False,
    ) -> None:
        if setting_id is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Serial prefix setting is required",
            )

        setting = self.product_setting_repo.get_by_id(setting_id, organization_id)
        if not setting:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Serial prefix setting not found",
            )
        if setting.setting_type != "serial_prefix":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Selected setting is not a serial prefix setting",
            )
        if not setting.is_active and not allow_inactive:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Selected serial prefix setting is inactive",
            )

    def _validate_optional_block_setting(
        self,
        setting_id: UUID | None,
        expected_type: str,
        organization_id: UUID,
    ) -> None:
        if setting_id is None:
            return
        setting = self.product_setting_repo.get_by_id(setting_id, organization_id)
        if not setting:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"{expected_type.title()} setting not found",
            )
        if setting.setting_type != expected_type:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Selected setting is not a {expected_type} setting",
            )
        if not setting.is_active:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Selected {expected_type} setting is inactive",
            )

    def create_product(
        self, data: QRProductCreate, organization_id: UUID, user_id: UUID
    ) -> QRProduct:
        product_dict = data.model_dump()
        packaging_details = product_dict.pop("packaging_details", None)
        if packaging_details is not None:
            extra = dict(product_dict.get("extra_data") or {})
            # packaging_details contains Decimal values; the JSONB serializer
            # (json.dumps) can't handle Decimal, so convert to JSON-safe floats.
            extra["packaging_details"] = {
                key: (float(value) if isinstance(value, Decimal) else value)
                for key, value in packaging_details.items()
            }
            product_dict["extra_data"] = extra
        product_dict["organization_id"] = organization_id
        product_dict["created_by"] = user_id
        product_dict["updated_by"] = user_id
        self._validate_shelf_life_setting(
            product_dict.get("shelf_life_setting_id"), organization_id
        )
        self._validate_serial_prefix_setting(
            product_dict.get("serial_prefix_setting_id"), organization_id
        )

        # Validate brand_id belongs to the organization when provided
        if product_dict.get("brand_id"):
            from app.repositories.brand_repository import BrandRepository

            brand_repo = BrandRepository(self.db)
            brand = brand_repo.get_by_id(product_dict["brand_id"], organization_id)
            if not brand:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Brand not found",
                )

        qr_product = self.product_repo.create(product_dict)

        # Auto-create a corresponding inventory Item linked to this QR product.
        # This ensures every QR product has a trackable item in the ERP without
        # requiring a separate frontend call.
        self._create_linked_item(
            qr_product, organization_id, user_id, packaging_details
        )

        # NOTE: field-level sync (product_item_sync_service) was removed in Phase 4.
        self.db.refresh(qr_product)

        return qr_product

    def _create_linked_item(  # noqa: C901
        self,
        qr_product: QRProduct,
        organization_id: UUID,
        user_id: UUID,
        packaging_details: dict | None = None,
    ) -> None:
        """Create an inventory Item that references this QR product.

        Uses the QR product's name as the item name. The item_code is
        auto-generated via DocumentNumberingService. Errors are logged but
        never bubble up — a failed item creation must not roll back the
        QR product itself.
        """
        item_code = None
        try:
            from app.models.base import ItemStatus, ItemType
            from app.models.item import Item
            from app.services.document_numbering_service import DocumentNumberingService

            # Guard: if an Item already exists referencing this QR product, skip creation.
            existing = (
                self.db.query(Item)
                .filter(Item.qr_product_id == qr_product.id, Item.deleted_at.is_(None))
                .first()
            )
            if existing:
                logger.info(
                    "Linked item already exists for QR product '%s' (product_id=%s) — skipping auto-create. Existing item: %s",
                    qr_product.name,
                    qr_product.id,
                    existing.item_code,
                )
                return

            item_code = DocumentNumberingService(self.db).get_next_number(
                organization_id, "item"
            )

            item = Item(
                organization_id=organization_id,
                item_code=item_code,
                item_name=qr_product.name,
                description=qr_product.generic_name,
                item_type=ItemType.STOCK,
                uom="Nos",
                sku=qr_product.sku or qr_product.gtin,
                gtin=qr_product.gtin,
                brand_id=qr_product.brand_id,
                maintain_stock=True,
                status=ItemStatus.ACTIVE,
                qr_product_id=qr_product.id,
                image_url=qr_product.image_url,
                created_by=user_id,
                updated_by=user_id,
            )
            self.db.add(item)
            self.db.commit()
            self.db.refresh(item)
            logger.info(
                "Auto-created item '%s' (id=%s) linked to QR product '%s' (id=%s)",
                item.item_code,
                item.id,
                qr_product.name,
                qr_product.id,
            )

            if packaging_details:
                try:
                    from app.schemas.item import ItemPackagingDetails
                    from app.services.item_service import ItemService

                    details = ItemPackagingDetails(
                        unit_name=packaging_details.get("unit_name") or "Each",
                        conversion_factor=packaging_details.get("conversion_factor")
                        or Decimal("1"),
                        items_per_master_pack=packaging_details.get(
                            "items_per_master_pack"
                        ),
                        length_mm=packaging_details.get("length_mm"),
                        width_mm=packaging_details.get("width_mm"),
                        height_mm=packaging_details.get("height_mm"),
                        weight_grams=packaging_details.get("weight_grams"),
                    )
                    ItemService(self.db)._upsert_base_packaging_unit(
                        item, details, organization_id
                    )
                    self.db.commit()
                except Exception as exc:
                    self.db.rollback()
                    logger.warning(
                        "Failed to upsert packaging unit for linked item '%s': %s",
                        item.id,
                        exc,
                    )
        except Exception as exc:
            # Roll back only the item insert, keep the QR product committed
            self.db.rollback()

            # Normalize detection of unique-constraint / integrity failures.
            # Some DB drivers wrap the underlying IntegrityError (DBAPIError),
            # so inspect __cause__ / orig when available.
            from sqlalchemy.exc import DBAPIError, IntegrityError

            def _is_integrity_error(e: BaseException) -> bool:
                if isinstance(e, IntegrityError):
                    return True
                if isinstance(e, DBAPIError):
                    # DBAPIError may wrap an underlying DB-API error in .orig
                    orig = getattr(e, "orig", None)
                    if orig and "unique" in str(orig).lower():
                        return True
                # Inspect chained exception
                cause = getattr(e, "__cause__", None)
                if cause:
                    return _is_integrity_error(cause)
                return False

            try:
                if _is_integrity_error(exc):
                    logger.warning(
                        "IntegrityError while creating item for QR product '%s' (id=%s): %s — attempting to attach to existing item",
                        qr_product.name,
                        qr_product.id,
                        exc,
                    )

                    existing_by_code = None
                    if item_code:
                        existing_by_code = (
                            self.db.query(Item)
                            .filter(
                                Item.organization_id == organization_id,
                                Item.item_code == item_code,
                                Item.deleted_at.is_(None),
                            )
                            .first()
                        )
                    if existing_by_code:
                        # Link existing item to this QR product if not already linked
                        if existing_by_code.qr_product_id != qr_product.id:
                            existing_by_code.qr_product_id = qr_product.id
                            existing_by_code.updated_by = user_id
                            self.db.add(existing_by_code)
                            self.db.commit()
                            logger.info(
                                "Attached existing item '%s' (id=%s) to QR product '%s' (id=%s)",
                                existing_by_code.item_code,
                                existing_by_code.id,
                                qr_product.name,
                                qr_product.id,
                            )
                            return
                        else:
                            logger.info(
                                "Existing item '%s' (id=%s) already linked to QR product '%s'",
                                existing_by_code.item_code,
                                existing_by_code.id,
                                qr_product.id,
                            )
                            return
            except Exception as exc2:
                # If attachment attempt failed, fall through to error log below
                logger.error(
                    "Failed to attach existing item for QR product '%s' (id=%s): %s",
                    qr_product.name,
                    qr_product.id,
                    exc2,
                )

            logger.error(
                "Failed to auto-create item for QR product '%s' (id=%s): %s",
                qr_product.name,
                qr_product.id,
                exc,
            )

    def get_product(self, product_id: UUID, organization_id: UUID) -> QRProduct:
        product = self.product_repo.get_by_id(product_id, organization_id)
        if not product:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="QR product not found"
            )
        return product

    def list_products(
        self,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
        is_active: bool | None = None,
    ) -> tuple[list[QRProduct], dict]:
        items, total = self.product_repo.list(
            organization_id, page, page_size, search, is_active
        )
        total_pages = max(1, (total + page_size - 1) // page_size)
        pagination = {
            "page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
        }
        return items, pagination

    def update_product(
        self,
        product_id: UUID,
        data: QRProductUpdate,
        organization_id: UUID,
        user_id: UUID,
    ) -> QRProduct:
        product = self.get_product(product_id, organization_id)
        update_dict = data.model_dump(exclude_unset=True)
        packaging_details = update_dict.pop("packaging_details", None)
        if packaging_details is not None:
            extra = dict(update_dict.get("extra_data") or {})
            extra["packaging_details"] = packaging_details
            update_dict["extra_data"] = extra

        if "shelf_life_setting_id" in update_dict:
            self._validate_shelf_life_setting(
                update_dict["shelf_life_setting_id"],
                organization_id,
                allow_inactive=(
                    update_dict["shelf_life_setting_id"]
                    == getattr(product, "shelf_life_setting_id", None)
                ),
            )
        if "serial_prefix_setting_id" in update_dict:
            self._validate_serial_prefix_setting(
                update_dict["serial_prefix_setting_id"],
                organization_id,
                allow_inactive=(
                    update_dict["serial_prefix_setting_id"]
                    == getattr(product, "serial_prefix_setting_id", None)
                ),
            )

        # brand_id is immutable after creation
        if "brand_id" in update_dict:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="brand_id cannot be modified after creation",
            )

        update_dict["updated_by"] = user_id
        product = self.product_repo.update(product, update_dict)

        if packaging_details is not None:
            try:
                from app.models.item import Item
                from app.schemas.item import ItemPackagingDetails
                from app.services.item_service import ItemService

                linked_item = (
                    self.db.query(Item)
                    .filter(
                        Item.qr_product_id == product.id,
                        Item.deleted_at.is_(None),
                    )
                    .first()
                )
                if linked_item is not None:
                    details = ItemPackagingDetails(
                        unit_name=packaging_details.get("unit_name") or "Each",
                        conversion_factor=packaging_details.get("conversion_factor")
                        or Decimal("1"),
                        items_per_master_pack=packaging_details.get(
                            "items_per_master_pack"
                        ),
                        length_mm=packaging_details.get("length_mm"),
                        width_mm=packaging_details.get("width_mm"),
                        height_mm=packaging_details.get("height_mm"),
                        weight_grams=packaging_details.get("weight_grams"),
                    )
                    ItemService(self.db)._upsert_base_packaging_unit(
                        linked_item, details, organization_id
                    )
                    self.db.commit()
            except Exception as exc:
                logger.error("Failed to upsert linked item packaging: %s", exc)

        # NOTE: field-level sync (product_item_sync_service) was removed in Phase 4.
        return product

    def delete_product(
        self, product_id: UUID, organization_id: UUID, user_id: UUID
    ) -> None:
        product = self.get_product(product_id, organization_id)
        self.product_repo.soft_delete(product, user_id)

    def update_product_image(
        self,
        product_id: UUID,
        image_type: str,
        image_url: str | None,
        organization_id: UUID,
        user_id: UUID,
    ) -> tuple[QRProduct, str | None]:
        product = self.get_product(product_id, organization_id)
        field = {
            "logo": "image_url",
            "banner": "banner_image_url",
        }.get(image_type)
        if field is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="image_type must be 'logo' or 'banner'",
            )

        previous_url = getattr(product, field)
        updated = self.product_repo.update(
            product,
            {
                field: image_url,
                "updated_by": user_id,
            },
        )
        return updated, previous_url

    # ── QR Blocks ─────────────────────────────────────────────────────────────

    def generate_block(
        self,
        product_id: UUID,
        data: QRBlockCreate,
        organization_id: UUID,
        user_id: UUID,
        org_credit_limit: int = 0,
    ) -> QRBlock:
        """Generate synchronously for internal callers and focused tests."""
        block = self.create_block_job(
            product_id,
            data,
            organization_id,
            user_id,
        )
        return self.process_block(
            block.id,
            organization_id,
            _claimed_block=block,
        )

    def create_block_job(
        self,
        product_id: UUID,
        data: QRBlockCreate,
        organization_id: UUID,
        user_id: UUID,
    ) -> QRBlock:
        """Validate a request, create its pending Block, and reserve credits."""
        product = self.get_product(product_id, organization_id)
        self._validate_optional_block_setting(
            data.channel_setting_id, "channel", organization_id
        )
        self._validate_optional_block_setting(
            data.destination_setting_id, "destination", organization_id
        )

        if self.block_repo.batch_exists(data.batch, organization_id):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Batch '{data.batch}' already exists",
            )

        self._validate_batch_product_mapping(
            data.batch,
            product_id,
            organization_id,
        )

        sku = self._get_block_sku(
            data.sku_id,
            product_id,
            organization_id,
        )

        try:
            qr_type = normalize_qr_type(data.qr_type or product.qr_type)
            serial_type = normalize_serial_number_type(
                (sku.sr_number_type if sku else None) or product.sr_number_type
            )
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=str(exc),
            ) from exc

        qr_type = qr_type or QRType.DYNAMIC
        if serial_type is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    "Product serial number type is not configured. "
                    "Update the product before generating a block."
                ),
            )
        serial_prefix = product.serial_prefix
        if not serial_prefix:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    "Product serial prefix is not configured. "
                    "Update the product before generating a block."
                ),
            )
        self._validate_block_generation_options(
            qr_type,
            serial_type,
            data.quantity,
            data.starting_serial,
        )
        requested_serials = self._build_deterministic_serials(
            qr_type,
            serial_type,
            serial_prefix,
            data.batch,
            data.starting_serial,
            data.quantity,
        )
        self._ensure_serials_available(requested_serials)

        # The balance is checked again under a row lock while reserving.
        self.credit_service.check_balance(organization_id, data.quantity)

        block_dict = data.model_dump()
        block_dict["qr_type"] = qr_type.value
        block_dict["sr_number_type"] = serial_type.value
        block_dict["serial_prefix"] = serial_prefix
        block_dict["product_id"] = product_id
        block_dict["organization_id"] = organization_id
        block_dict["created_by"] = user_id
        block_dict["updated_by"] = user_id
        block_dict["status"] = "pending"
        block_dict["task_status"] = "pending"
        block_dict["generated_count"] = 0
        block_dict["progress"] = 0

        try:
            block = self.block_repo.create(block_dict, commit=False)
            # Repository implementations return a fully populated ORM object.
            # Keeping these assignments explicit also makes the processing
            # contract clear for alternate repository implementations.
            for field, value in block_dict.items():
                if not hasattr(block, field):
                    setattr(block, field, value)
            self.credit_service.reserve_credits(
                organization_id,
                block.id,
                data.quantity,
                commit=False,
            )
            self.db.commit()
            self.db.refresh(block)
        except IntegrityError as exc:
            self.db.rollback()
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Batch '{data.batch}' already exists",
            ) from exc
        return block

    def _validate_batch_product_mapping(
        self,
        batch_no: str,
        product_id: UUID,
        organization_id: UUID,
    ) -> None:
        """Reject a block whose batch label already belongs to another product.

        QR block batch names are free-form labels, but when the same label has
        already been created as a WMS batch for an item that is linked to a
        different QR product, generating codes under this product would cross
        product boundaries.
        """
        from app.models.batch import Batch
        from app.models.item import Item

        linked_batches = (
            self.db.query(Batch)
            .join(Item, Item.id == Batch.item_id)
            .filter(
                Batch.organization_id == organization_id,
                Batch.batch_no == batch_no,
            )
            .all()
        )
        for batch in linked_batches:
            item = batch.item
            if item is None or item.qr_product_id is None:
                # Not linked to a QR product — nothing to compare against.
                continue
            if item.qr_product_id == product_id:
                continue

            other_product = self.db.get(QRProduct, item.qr_product_id)
            other_name = (
                other_product.name if other_product else str(item.qr_product_id)
            )
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"Batch '{batch_no}' belongs to product '{other_name}'. "
                    "Select a batch for the product you are generating codes for."
                ),
            )

    def _get_block_sku(
        self,
        sku_id: UUID | None,
        product_id: UUID,
        organization_id: UUID,
    ):
        if sku_id is None:
            return None
        sku = self.sku_repo.get_by_id(sku_id, organization_id)
        if not sku or sku.product_id != product_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Product SKU not found",
            )
        if not sku.is_active:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Product SKU is inactive",
            )
        return sku

    def assign_block_task(
        self,
        block_id: UUID,
        organization_id: UUID,
        task_id: str,
    ) -> QRBlock:
        block = self.block_repo.get_by_id(block_id, organization_id)
        if block is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="QR Block not found",
            )
        if block.status not in {"pending", "failed"}:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Block cannot be queued from status '{block.status}'",
            )
        block.task_id = task_id
        block.status = "pending"
        block.task_status = "pending"
        block.progress = 0
        block.generated_count = 0
        block.error_code = None
        block.error_message = None
        block.completed_at = None
        self.db.commit()
        self.db.refresh(block)
        return block

    def fail_block_enqueue(
        self,
        block_id: UUID,
        organization_id: UUID,
    ) -> None:
        block = self.block_repo.get_by_id(block_id, organization_id)
        if block is None or block.status == "completed":
            return
        block.status = "failed"
        block.task_status = "failed"
        block.error_code = "queue_unavailable"
        block.error_message = "QR generation queue is unavailable"
        block.task_id = None
        self.credit_service.release_reserved_credits(organization_id, block_id)

    def fail_block_processing(
        self,
        block_id: UUID,
        organization_id: UUID,
    ) -> None:
        """Make a task-level worker failure visible and return held credits."""
        self.db.rollback()
        block = self.block_repo.get_by_id(block_id, organization_id)
        if block is None or block.status == "completed":
            return
        block.status = "failed"
        block.task_status = "failed"
        block.error_code = "worker_failed"
        block.error_message = "QR generation worker failed"
        block.task_id = None
        block.progress = 0
        block.generated_count = 0
        self.credit_service.release_reserved_credits(organization_id, block_id)
        self.db.commit()

    def retry_block_job(
        self,
        block_id: UUID,
        organization_id: UUID,
    ) -> QRBlock:
        """Reset a failed Block and reserve its credits for a new worker task."""
        block = self.block_repo.get_by_id_for_update(block_id, organization_id)
        if block is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="QR Block not found",
            )
        if block.status != "failed":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Only failed QR Blocks can be retried",
            )
        self.credit_service.reserve_credits(
            organization_id,
            block.id,
            block.quantity,
            commit=False,
        )
        block.status = "pending"
        block.task_status = "pending"
        block.task_id = None
        block.progress = 0
        block.generated_count = 0
        block.error_code = None
        block.error_message = None
        block.completed_at = None
        self.db.commit()
        self.db.refresh(block)
        return block

    def process_block(
        self,
        block_id: UUID,
        organization_id: UUID,
        task_id: str | None = None,
        _claimed_block: QRBlock | None = None,
    ) -> QRBlock:
        """Idempotently process one queued Block within a worker session."""
        block = _claimed_block or self.block_repo.get_by_id_for_update(
            block_id,
            organization_id,
        )
        if block is None:
            raise RuntimeError("Queued QR Block no longer exists")
        if block.status == "completed":
            return block
        if block.status == "failed":
            return block
        if task_id and block.task_id != task_id:
            logger.warning(
                "Ignoring stale QR Block task: block_id=%s task_id=%s current=%s",
                block.id,
                task_id,
                block.task_id,
            )
            return block

        product = self.get_product(block.product_id, organization_id)

        # A redelivered task may follow a worker crash after item insertion.
        # Remove only this tenant's prior active rows before regenerating.
        if block.status == "in_progress":
            self.item_repo.soft_delete_by_block(block.id, organization_id)

        block.status = "in_progress"
        block.task_status = "in_progress"
        block.progress = 5
        block.error_code = None
        block.error_message = None
        self.db.commit()

        uploaded_artifact_key: str | None = None
        try:
            generated_items = self._generate_product_items(
                block,
                product,
                organization_id,
                block.created_by,
            )
            block = self.block_repo.get_by_id(block.id, organization_id) or block
            block.generated_count = len(generated_items)
            block.progress = 70
            self.db.commit()

            uploaded_artifact_key = self._store_block_artifact(block, generated_items)
            block.progress = 90
            self.db.commit()

            block.status = "completed"
            block.task_status = "completed"
            block.completed_at = datetime.now(UTC)
            block.generated_count = block.quantity
            block.progress = 100
            self.credit_service.consume_reserved_credits(
                organization_id,
                block.id,
                block.created_by,
            )

        except Exception as exc:
            self.db.rollback()
            if uploaded_artifact_key:
                try:
                    from app.services import storage_service

                    storage_service.delete_qr_artifact(uploaded_artifact_key)
                except Exception:
                    logger.exception(
                        "Failed to clean up QR artifact after generation failure: "
                        "block_id=%s object_key=%s",
                        block.id,
                        uploaded_artifact_key,
                    )
            self.item_repo.soft_delete_by_block(block.id, organization_id)
            failed_block = self.block_repo.get_by_id(block.id, organization_id) or block
            failed_block.status = "failed"
            failed_block.task_status = "failed"
            failed_block.error_code = "generation_failed"
            failed_block.error_message = "QR block generation failed"
            failed_block.progress = 0
            failed_block.generated_count = 0
            failed_block.artifact_object_key = None
            failed_block.artifact_size_bytes = None
            failed_block.artifact_checksum_sha256 = None
            failed_block.artifact_generated_at = None
            self.credit_service.release_reserved_credits(
                organization_id,
                block.id,
            )
            logger.exception(
                "Block generation failed: block_id=%s product_id=%s org=%s",
                block.id,
                block.product_id,
                organization_id,
            )
            if isinstance(exc, IntegrityError):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=(
                        "A generated serial number already exists; retry with "
                        "a different batch or serial range"
                    ),
                ) from exc
            raise

        logger.info(
            "QR block generated: block_id=%s product_id=%s qty=%d org=%s",
            block.id,
            block.product_id,
            block.quantity,
            organization_id,
        )
        return block

    # ── Serial number helpers ─────────────────────────────────────────────────

    @staticmethod
    def _build_deterministic_serials(
        qr_type: QRType,
        serial_type: SerialNumberType,
        prefix: str,
        batch: str,
        starting_serial: str | None,
        quantity: int,
    ) -> list[str]:
        def with_prefix(suffix: str) -> str:
            return f"{prefix}-{suffix}" if prefix else suffix

        if qr_type == QRType.STATIC:
            return [with_prefix(batch)]
        if serial_type not in {SerialNumberType.S8DN, SerialNumberType.S10DN}:
            return []

        width = 8 if serial_type == SerialNumberType.S8DN else 10
        start = int(starting_serial or "1")
        return [
            with_prefix(f"{value:0{width}d}")
            for value in range(start, start + quantity)
        ]

    def _ensure_serials_available(
        self,
        serials: list[str],
    ) -> None:
        if not serials:
            return
        existing = self.item_repo.get_existing_serials_global(serials)
        if not existing:
            return

        preview = ", ".join(sorted(existing)[:5])
        remaining = len(existing) - 5
        suffix = f" and {remaining} more" if remaining > 0 else ""
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(f"Serial numbers already exist: {preview}{suffix}"),
        )

    @staticmethod
    def _validate_block_generation_options(
        qr_type: QRType,
        serial_type: SerialNumberType,
        quantity: int,
        starting_serial: str | None,
    ) -> list[dict]:
        if qr_type == QRType.STATIC and quantity != 1:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Static QR generation requires quantity=1",
            )
        if qr_type == QRType.STATIC:
            if starting_serial is not None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="starting_serial is not valid for Static QR",
                )
            return

        if serial_type not in {SerialNumberType.S8DN, SerialNumberType.S10DN}:
            if starting_serial is not None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=(
                        "starting_serial is only valid for sequential serial numbers"
                    ),
                )
            return

        if starting_serial is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="starting_serial is required for sequential serial numbers",
            )
        if not starting_serial.isdigit():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="starting_serial must contain digits only",
            )

        max_value = (
            99_999_999 if serial_type == SerialNumberType.S8DN else 9_999_999_999
        )
        if int(starting_serial) + quantity - 1 > max_value:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Sequential range exceeds {serial_type.value} capacity",
            )

    def _get_serial_generator(
        self, sr_number_type: str | None, starting_serial: str | None = None
    ):
        """Return a callable that produces serial numbers for the given type."""
        sr_type = (sr_number_type or "R6DAN").upper()
        if sr_type == "R8DAN":
            return generate_r8dan
        if sr_type == "R4DAN":
            return generate_r4dan
        if sr_type == "S8DN":
            gen = sequential_s8dn(int(starting_serial or "1"))
            return lambda: next(gen)
        if sr_type == "S10DN":
            gen = sequential_s10dn(int(starting_serial or "1"))
            return lambda: next(gen)
        # Default: R6DAN
        return generate_r6dan

    @staticmethod
    def _generate_secret_code() -> str:
        """Generate a 12-character alphanumeric secret code for SecureCode QR."""
        return "".join(
            secrets.choice(string.ascii_uppercase + string.digits) for _ in range(12)
        )

    # ── Item generation ───────────────────────────────────────────────────────

    def _generate_product_items(  # noqa: C901
        self,
        block: QRBlock,
        product: QRProduct,
        organization_id: UUID,
        user_id: UUID,
    ) -> None:
        """Generate ProductItem rows for the block.

        Handles brand-based signing and QR type-specific behaviour:
        - S (Static): same serial for all items
        - SC (SecureCode): 12-char secret per item
        - O (OneTime): qr_active based on activation_method
        - B (Dual): two URLs per item (covert + overt)
        - D (Dynamic) / default: unique URL per item
        """
        now = datetime.now(UTC)
        qr_type = normalize_qr_type(block.qr_type) or QRType.DYNAMIC
        sr_number_type = block.sr_number_type or product.sr_number_type
        prefix = block.serial_prefix or ""

        # Determine if we need signing (product linked to a brand)
        brand = None
        private_key = None
        gtin = product.gtin or ""

        if block.sku_id:
            sku = self.sku_repo.get_by_id(block.sku_id, organization_id)
            if sku and sku.gtin:
                gtin = sku.gtin

        if not gtin:
            raise RuntimeError("Product GTIN is required for QR generation")

        if product.brand_id and self.key_service:
            from app.repositories.brand_repository import BrandRepository

            brand_repo = BrandRepository(self.db)
            brand = brand_repo.get_by_id(product.brand_id, organization_id)
            if brand and brand.private_key_encrypted:
                private_key = self.key_service.decrypt_private_key(
                    brand.private_key_encrypted
                )

        serial_gen = self._get_serial_generator(sr_number_type, block.starting_serial)

        def with_prefix(suffix: str) -> str:
            return f"{prefix}-{suffix}" if prefix else suffix

        normalized_serial_type = normalize_serial_number_type(sr_number_type)
        serials = self._build_deterministic_serials(
            qr_type,
            normalized_serial_type or SerialNumberType.R6DAN,
            prefix,
            block.batch,
            block.starting_serial,
            block.quantity,
        )
        if serials:
            self._ensure_serials_available(serials)
        else:
            attempts = 0
            while len(serials) < block.quantity:
                missing = block.quantity - len(serials)
                candidates = [with_prefix(serial_gen()) for _ in range(missing)]
                candidates = list(dict.fromkeys(candidates))
                existing = self.item_repo.get_existing_serials_global(candidates)
                serials.extend(
                    candidate
                    for candidate in candidates
                    if candidate not in existing and candidate not in serials
                )
                attempts += 1
                if attempts >= 20 and len(serials) < block.quantity:
                    raise RuntimeError(
                        "Unable to generate the requested number of unique serials"
                    )

        items: list[dict] = []

        for serial in serials:
            item_dict: dict = {
                "id": uuid.uuid4(),
                "organization_id": organization_id,
                "product_id": block.product_id,
                "block_id": block.id,
                "sku_id": block.sku_id,
                "serial_number": serial,
                "created_by": user_id,
                "updated_by": user_id,
                "created_at": now,
                "updated_at": now,
            }

            # SecureCode: generate 12-char secret
            if qr_type == QRType.SECURE_CODE:
                item_dict["secrete_code"] = self._generate_secret_code()

            is_pre_activated = product.activation_method != "post"
            item_dict["qr_active"] = is_pre_activated
            item_dict["qr_deactive"] = not is_pre_activated
            item_dict["qr_deactive_unit"] = not is_pre_activated

            # Sign if brand is linked
            if private_key is not None:
                sig, ts = sign_qr_item(self.key_service, private_key, serial)
                url = build_qr_url(
                    settings.qr_domain,
                    gtin,
                    serial,
                    ts,
                    sig,
                    base_url=settings.qr_base_url,
                )
                # Dual QR: retain distinct overt and covert signed URLs.
                if qr_type == QRType.DUAL:
                    sig2, ts2 = sign_qr_item(self.key_service, private_key, serial)
                    covert_url = (
                        build_qr_url(
                            settings.qr_domain,
                            gtin,
                            serial,
                            ts2,
                            sig2,
                            base_url=settings.qr_base_url,
                        )
                        + "&qr=covert"
                    )
                    overt_url = f"{url}&qr=overt"
                    short_overt_url = self.qr_shortener.shorten(overt_url)
                    short_covert_url = self.qr_shortener.shorten(covert_url)
                    item_dict["token_id"] = short_overt_url
                    item_dict["extra_data"] = {
                        "long_url": overt_url,
                        "short_url": short_overt_url,
                        "overt_url": short_overt_url,
                        "covert_url": short_covert_url,
                        "overt_long_url": overt_url,
                        "covert_long_url": covert_url,
                    }
                else:
                    short_url = self.qr_shortener.shorten(url)
                    item_dict["token_id"] = short_url
                    item_dict["extra_data"] = {
                        "long_url": url,
                        "short_url": short_url,
                    }
            else:
                # No brand/key — still generate a URL for Excel download
                base = settings.qr_base_url or f"https://{settings.qr_domain}"
                item_dict["token_id"] = f"{base}/g/{gtin}/s/{serial}"

            items.append(item_dict)

        self.item_repo.bulk_create(items)

        # Create QSeal parent nodes after the child items are persisted.
        master_pack_enabled = getattr(block, "master_pack_enabled", False)
        master_pack_size = getattr(block, "master_pack_size", None)
        if master_pack_enabled and master_pack_size and master_pack_size > 0:
            self._create_qseal_parents(block, items, organization_id, user_id, now)

        return items

    @staticmethod
    def _items_to_excel_rows(items: list) -> list[dict]:
        rows = []
        for item in items:
            if isinstance(item, dict):
                serial = item.get("serial_number")
                primary_url = item.get("token_id") or ""
                secret_code = item.get("secrete_code") or ""
                extra_data = item.get("extra_data") or {}
            else:
                serial = item.serial_number
                primary_url = item.token_id or ""
                secret_code = item.secrete_code or ""
                extra_data = item.extra_data or {}
            rows.append(
                {
                    "serial": serial,
                    "primary_url": primary_url,
                    "secret_code": secret_code,
                    "overt_url": extra_data.get("overt_url", primary_url),
                    "covert_url": extra_data.get("covert_url", ""),
                }
            )
        return rows

    def _store_block_artifact(
        self,
        block: QRBlock,
        generated_items: list[dict],
    ) -> str | None:
        """Build and store the completed workbook.

        Stores to private S3 when configured. Otherwise returns ``None`` — the
        workbook is regenerated on demand from ProductItems at download time
        and cached on the local volume (core-service-volume).
        """
        from app.services import storage_service

        if not settings.aws_s3_bucket:
            return None

        rows = self._items_to_excel_rows(generated_items)
        excel_bytes = _build_excel(rows, block.qr_type, bool(block.qr_image))
        object_key = storage_service.build_qr_artifact_key(
            block.organization_id,
            block.product_id,
            block.id,
        )
        filename = f"qr_block_{block.id}.xlsx"
        storage_service.store_qr_artifact(excel_bytes, object_key, filename)
        block.artifact_object_key = object_key
        block.artifact_size_bytes = len(excel_bytes)
        block.artifact_checksum_sha256 = hashlib.sha256(excel_bytes).hexdigest()
        block.artifact_generated_at = datetime.now(UTC)
        return object_key

    def _create_qseal_parents(
        self,
        block: QRBlock,
        items: list[dict],
        organization_id: UUID,
        user_id: UUID,
        now: datetime,
    ) -> None:
        """Create QSealTrack parent nodes for master packs.

        Groups ProductItems into chunks of master_pack_size.
        For each chunk, creates a QSealTrack (shipper) parent and
        QSealParameters entries linking items to their parent.
        """
        from app.models.qseal import QSealParameters, QSealTrack
        from app.repositories.qseal_repository import QSealRepository

        qseal_repo = QSealRepository(self.db)
        pack_size = block.master_pack_size
        parent_count = 0

        for chunk_start in range(0, len(items), pack_size):
            chunk = items[chunk_start : chunk_start + pack_size]
            if not chunk:
                continue

            # Create QSealTrack parent (shipper level)
            parent_serial = qseal_repo.generate_serial(prefix="QSL")
            # Truncate name to fit VARCHAR(20)
            parent_name = f"MP-{block.batch[:10]}-{parent_count + 1}"[:20]
            parent_node = QSealTrack(
                id=uuid.uuid4(),
                organization_id=organization_id,
                qseal_type="shipper",
                name=parent_name,
                capacity=pack_size,
                serial_number=parent_serial,
                qseal_code_link=f"/qseal/{parent_serial}",
                app_cascade_map=False,
                created_at=now,
            )
            self.db.add(parent_node)
            self.db.flush()  # Get the parent ID

            # Create QSealParameters for each item in the chunk
            for item in chunk:
                qsp = QSealParameters(
                    id=uuid.uuid4(),
                    organization_id=organization_id,
                    product_id=block.product_id,
                    block_id=block.id,
                    serial_number=item["serial_number"],
                    manufacturing_date=block.manufacture_date or now.date(),
                    expiry_date=block.expiry_date or now.date(),
                    manufacturing_unit="",
                    dispatch_batch=block.batch,
                    batch_size=pack_size,
                    qseal_settings=False,
                    qseal_cascade=False,
                    parent_id=parent_node.id,
                    extra_data={
                        "item_id": str(item["id"]),
                        "master_pack_index": parent_count + 1,
                    },
                    created_by=user_id,
                    created_at=now,
                )
                self.db.add(qsp)

            parent_count += 1
            logger.info(
                "[QSEAL] master-pack parent created serial=%s block=%s org=%s items=%d",
                parent_serial,
                block.id,
                organization_id,
                len(chunk),
            )

        self.db.commit()

        # Store parent count on block for reference
        block.extra_data = (block.extra_data or {}) | {
            "qseal_parent_count": parent_count
        }
        self.db.commit()

        logger.info(
            "[QSEAL] master-pack complete block=%s parents=%d total_items=%d",
            block.id,
            parent_count,
            len(items),
        )

    def get_block(self, block_id: UUID, organization_id: UUID) -> QRBlock:
        block = self.block_repo.get_by_id(block_id, organization_id)
        if not block:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="QR block not found"
            )
        return block

    def get_block_detail(self, block_id: UUID, organization_id: UUID) -> QRBlock:
        """Return a tenant-scoped Block with its current activation summary."""
        block = self.get_block(block_id, organization_id)
        total, active = self.item_repo.get_activation_summary(block_id, organization_id)
        if total and active == total:
            activation_status = "activated"
        elif active:
            activation_status = "partially_activated"
        else:
            activation_status = "deactivated"

        block.activation_status = activation_status
        block.activated_count = active
        block.deactivated_count = total - active
        return block

    def get_block_download_url(
        self, block_id: UUID, organization_id: UUID
    ) -> tuple[str, datetime]:
        """Return a signed download URL for a completed block's Excel file."""
        from app.services import storage_service

        block = self.get_block(block_id, organization_id)

        if block.status != "completed":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Block is not ready (status: {block.status})",
            )

        if block.artifact_object_key:
            try:
                return storage_service.get_qr_artifact_signed_url(
                    block.artifact_object_key
                )
            except Exception as exc:
                logger.exception(
                    "Failed to create QR artifact download URL: "
                    "block_id=%s organization_id=%s",
                    block.id,
                    organization_id,
                )
                raise HTTPException(
                    status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                    detail="QR artifact storage is temporarily unavailable",
                ) from exc

        expiry_minutes = 60
        expires_at = datetime.now(UTC) + timedelta(minutes=expiry_minutes)

        # If a GCS URL is stored, return it (signed or direct)
        if block.download_url:
            if storage_service.is_full_url(block.download_url):
                return block.download_url, expires_at
            signed_url = storage_service.get_signed_url(
                block.download_url, expiry_minutes
            )
            return signed_url, expires_at

        # No stored URL — raise so the endpoint falls back to streaming
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Download file not available",
        )

    def get_block_local_artifact_path(
        self,
        block_id: UUID,
        organization_id: UUID,
    ):
        """Return the local volume path used to cache a block's workbook."""
        from app.services import storage_service

        block = self.get_block(block_id, organization_id)
        if block.status != "completed":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Block is not ready (status: {block.status})",
            )
        return storage_service.build_qr_artifact_local_path(
            block.organization_id,
            block.product_id,
            block.id,
        )

    def get_block_excel_stream(
        self, block_id: UUID, organization_id: UUID
    ) -> tuple[bytes, str]:
        """
        Generate the Excel file on-demand from the block's ProductItems.
        Used when download_url is not set (GCS not configured).
        Returns (excel_bytes, filename).
        """
        from app.models.product_item import ProductItem as PIModel

        block = self.get_block(block_id, organization_id)

        if block.status != "completed":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Block is not ready (status: {block.status})",
            )

        qr_type = block.qr_type or QRType.DYNAMIC.value

        items = (
            self.db.query(PIModel)
            .filter(
                PIModel.block_id == block_id,
                PIModel.organization_id == organization_id,
                PIModel.deleted_at.is_(None),
            )
            .order_by(PIModel.created_at.asc())
            .all()
        )

        rows = self._items_to_excel_rows(items)
        excel_bytes = _build_excel(rows, qr_type, bool(block.qr_image))
        filename = f"qr_block_{block.batch}_{block_id}.xlsx"
        return excel_bytes, filename

    def list_blocks(
        self,
        product_id: UUID,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[QRBlock], dict]:
        # Validate product
        self.get_product(product_id, organization_id)
        items, total = self.block_repo.list_by_product(
            product_id, organization_id, page, page_size
        )
        total_pages = max(1, (total + page_size - 1) // page_size)
        pagination = {
            "page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
        }
        return items, pagination

    def list_blocks_by_org(
        self,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 20,
        block_status: str | None = None,
        product_id: UUID | None = None,
        search: str | None = None,
        qr_type: str | None = None,
        created_from: datetime | None = None,
        created_to: datetime | None = None,
    ) -> tuple[list[dict], dict]:
        for value, label in (
            (created_from, "created_from"),
            (created_to, "created_to"),
        ):
            if value is not None and value.utcoffset() is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"{label} must include a timezone offset",
                )
        if (
            created_from is not None
            and created_to is not None
            and created_from >= created_to
        ):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="created_from must be earlier than created_to",
            )

        rows, total = self.block_repo.list_by_org(
            organization_id,
            page,
            page_size,
            block_status,
            product_id,
            search.strip() if search else None,
            qr_type,
            created_from,
            created_to,
        )
        total_pages = max(1, (total + page_size - 1) // page_size)
        pagination = {
            "page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
        }
        enriched = []
        for block, product_name in rows:
            block_dict = {
                k: v for k, v in block.__dict__.items() if k != "_sa_instance_state"
            }
            block_dict["product_name"] = product_name
            block_dict["distribution_channel"] = block.distribution_channel
            block_dict["destination_market"] = block.destination_market
            block_dict["download_available"] = block.download_available
            enriched.append(block_dict)
        return enriched, pagination

    # ── Product Items ─────────────────────────────────────────────────────────

    def list_items(
        self,
        block_id: UUID,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[list[ProductItem], dict]:
        block = self.block_repo.get_by_id(block_id, organization_id)
        if not block:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="QR block not found"
            )
        items, total = self.item_repo.list_by_block(
            block_id, organization_id, page, page_size
        )
        total_pages = max(1, (total + page_size - 1) // page_size)
        pagination = {
            "page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
        }
        return items, pagination

    # ── QR Validate (public) ──────────────────────────────────────────────────

    def validate_qr(self, organization_id: UUID, req: QRValidateRequest) -> dict:
        """
        Authenticate a QR scan. Records the scan event and returns authenticity.
        This endpoint is typically called from the consumer-facing landing page.
        """
        item = self.item_repo.get_by_serial(req.serial_number, organization_id)
        if not item:
            return {
                "is_authentic": False,
                "is_suspicious": False,
                "scans": 0,
                "product_name": None,
                "message": "QR code not found. This product may be counterfeit.",
            }

        # Flag as suspicious if scanned more than once (configurable threshold)
        if item.scans >= 1 and not item.is_suspicious:
            item.is_suspicious = True
            self.db.flush()

        # Record scan event
        scan_data = {
            k: v
            for k, v in req.model_dump().items()
            if k != "serial_number" and v is not None
        }
        self.item_repo.record_scan(item, scan_data)

        product_name = item.product.name if item.product else None
        is_first_scan = item.scans == 1

        logger.info(
            "QR scan: serial=%s org=%s scans=%d suspicious=%s",
            req.serial_number,
            organization_id,
            item.scans,
            item.is_suspicious,
        )

        return {
            "is_authentic": True,
            "is_suspicious": item.is_suspicious,
            "scans": item.scans,
            "product_name": product_name,
            "message": (
                "Authentic product."
                if is_first_scan
                else f"Warning: This QR has been scanned {item.scans} times."
            ),
        }

    # ── QR Authenticate (public, ECDSA) ─────────────────────────────────────

    def authenticate(self, organization_id: UUID, data: AuthenticateRequest) -> dict:
        """
        Verify a QR scan using ECDSA signature verification.

        Requirements: 9.1-9.9, 8.4
        """
        # 1. Look up item by serial_number
        item = self.item_repo.get_by_serial(data.serial_number, organization_id)
        if not item:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Serial number not found",
            )

        # 2. Load product and brand via relationships
        product = item.product
        brand = product.brand if product else None

        # 3. Check post-activation: qr_active must be True
        if product and product.activation_method == "post" and not item.qr_active:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Product has not been activated",
            )

        # 4. Brand must exist for ECDSA verification
        if not brand:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No brand linked to product",
            )

        # 5. Reconstruct message and verify signature
        message = f"{data.serial_number}~{data.nonce}"
        is_valid = self.key_service.verify_signature(
            brand.public_key, message, data.cipher
        )

        if not is_valid:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "message": "Authentication Failed",
                    "authentic": False,
                },
            )

        # 6. Valid signature — update scan tracking
        item.scan_count = (item.scan_count or 0) + 1
        item.last_scanned_at = datetime.now(UTC)

        # 7. OneTime QR: deactivate after first successful verification
        if product.qr_type == "O":
            item.qr_active = False

        self.db.commit()

        logger.info(
            "QR authenticate: serial=%s org=%s authentic=True scan_count=%d",
            data.serial_number,
            organization_id,
            item.scan_count,
        )

        return {
            "message": "Authentic Product",
            "authentic": True,
            "product_name": product.name,
            "brand_name": brand.name,
            "gtin": product.gtin,
            "serial_number": data.serial_number,
        }

    # ── Scan Analytics ────────────────────────────────────────────────────────

    def get_scan_analytics(self, product_id: UUID, organization_id: UUID) -> dict:
        self.get_product(product_id, organization_id)
        return self.item_repo.get_scan_analytics(product_id, organization_id)

    # ── Activation Parameters ─────────────────────────────────────────────────

    def set_activation_params(
        self,
        data: QRActivationParamsCreate,
        organization_id: UUID,
        user_id: UUID,
    ):
        from app.models.qr_activation import QRActivationParameters

        params_dict = data.model_dump()
        params_dict["organization_id"] = organization_id
        params_dict["created_by"] = user_id
        params = QRActivationParameters(**params_dict)
        self.db.add(params)
        self.db.commit()
        self.db.refresh(params)
        return params
