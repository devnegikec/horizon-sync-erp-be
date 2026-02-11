"""Utilities and validators for bulk import/export operations"""

import csv
import io
import json
from enum import Enum
from typing import Any

import openpyxl
import pandas as pd
from pydantic import BaseModel, field_validator


class FileFormat(str, Enum):
    """Supported file formats"""

    CSV = "csv"
    XLSX = "xlsx"
    JSON = "json"


class BulkImportValidator:
    """Validator for bulk import data"""

    REQUIRED_FIELDS = {"item_code", "item_name"}
    VALID_COLUMNS = {
        "item_code", "item_name", "description", "item_group_id", "item_group_name",
        "item_type", "uom", "maintain_stock", "valuation_method",
        "allow_negative_stock", "has_variants", "variant_of", "has_batch_no",
        "has_serial_no", "batch_number_series", "serial_number_series",
        "standard_rate", "valuation_rate", "enable_auto_reorder",
        "reorder_level", "reorder_qty", "min_order_qty", "max_order_qty",
        "weight_per_unit", "weight_uom", "inspection_required_before_purchase",
        "inspection_required_before_delivery", "quality_inspection_template",
        "barcode", "status", "image_url", "tags", "custom_fields", "extra_data"
    }
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    MAX_ROWS = 10000

    @staticmethod
    def validate_columns(columns: list[str]) -> tuple[bool, list[str]]:
        """
        Validate if all columns in the file are valid for Item model.
        """
        invalid_columns = [col for col in columns if col not in BulkImportValidator.VALID_COLUMNS]
        if invalid_columns:
            return False, [f"Invalid columns found: {', '.join(invalid_columns)}"]
        return True, []

    @staticmethod
    def validate_file_size(file_size: int) -> tuple[bool, str | None]:
        """
        Validate file size.

        Args:
            file_size: File size in bytes

        Returns:
            Tuple of (is_valid, error_message)
        """
        if file_size > BulkImportValidator.MAX_FILE_SIZE:
            return False, f"File size exceeds {BulkImportValidator.MAX_FILE_SIZE / (1024 * 1024):.1f}MB limit"
        return True, None
    @staticmethod
    def validate_file_format(mime_type: str, file_format: str) -> tuple[bool, str | None]:
        """
        Validate file format.

        Args:
            mime_type: MIME type of the file
            file_format: File format (csv, xlsx, json)

        Returns:
            Tuple of (is_valid, error_message)
        """
        valid_formats = {
            "csv": ["text/csv", "application/csv"],
            "xlsx": ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
            "json": ["application/json"],
        }

        if file_format not in valid_formats:
            return False, f"Unsupported format: {file_format}"

        if mime_type not in valid_formats[file_format]:
            return False, f"Invalid MIME type {mime_type} for format {file_format}"

        return True, None

    @staticmethod
    def validate_row(row: dict, row_number: int) -> tuple[bool, list[str]]:
        """
        Validate a single row of import data.

        Args:
            row: Dictionary containing row data
            row_number: Row number for error reporting

        Returns:
            Tuple of (is_valid, list_of_error_messages)
        """
        errors = []

        # Check required fields
        for field in BulkImportValidator.REQUIRED_FIELDS:
            if field not in row or not row[field]:
                errors.append(f"Row {row_number}: Missing required field '{field}'")

        # Validate item_code format (should be alphanumeric with possible hyphens/underscores)
        if "item_code" in row and row["item_code"]:
            item_code = str(row["item_code"]).strip()
            if len(item_code) > 100:
                errors.append(f"Row {row_number}: item_code exceeds 100 characters")

        # Validate item_name length
        if "item_name" in row and row["item_name"]:
            item_name = str(row["item_name"]).strip()
            if len(item_name) > 255:
                errors.append(f"Row {row_number}: item_name exceeds 255 characters")

        return len(errors) == 0, errors


class FileParser:
    """Parser for different file formats"""

    @staticmethod
    def parse_csv(file_content: bytes) -> list[dict]:
        """
        Parse CSV file.

        Args:
            file_content: File content as bytes

        Returns:
            List of dictionaries representing rows
        """
        try:
            text = file_content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            rows = []
            for row in reader:
                # Strip whitespace from keys and values
                cleaned_row = {k.strip(): v.strip() if v else None for k, v in row.items()}
                rows.append(cleaned_row)
            return rows
        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {str(e)}")

    @staticmethod
    def parse_xlsx(file_content: bytes) -> list[dict]:
        """
        Parse XLSX file.

        Args:
            file_content: File content as bytes

        Returns:
            List of dictionaries representing rows
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    header = headers[col_idx] if col_idx < len(headers) else None
                    if header:
                        row_dict[str(header).strip()] = value
                if any(row_dict.values()):  # Skip empty rows
                    rows.append(row_dict)

            return rows
        except Exception as e:
            raise ValueError(f"Failed to parse XLSX: {str(e)}")

    @staticmethod
    def parse_json(file_content: bytes) -> list[dict]:
        """
        Parse JSON file.

        Args:
            file_content: File content as bytes

        Returns:
            List of dictionaries representing rows
        """
        try:
            data = json.loads(file_content.decode("utf-8"))
            if isinstance(data, list):
                return data
            elif isinstance(data, dict) and "items" in data:
                return data["items"]
            else:
                raise ValueError("JSON must be a list or contain 'items' key")
        except Exception as e:
            raise ValueError(f"Failed to parse JSON: {str(e)}")

    @staticmethod
    def parse_file(file_content: bytes, file_format: str) -> list[dict]:
        """
        Parse file based on format.

        Args:
            file_content: File content as bytes
            file_format: File format (csv, xlsx, json)

        Returns:
            List of dictionaries representing rows
        """
        if file_format == FileFormat.CSV:
            return FileParser.parse_csv(file_content)
        elif file_format == FileFormat.XLSX:
            return FileParser.parse_xlsx(file_content)
        elif file_format == FileFormat.JSON:
            return FileParser.parse_json(file_content)
        else:
            raise ValueError(f"Unsupported format: {file_format}")


class FileGenerator:
    """Generator for different export file formats"""

    @staticmethod
    def generate_csv(data: list[dict], headers: list[str] | None = None) -> bytes:
        """
        Generate CSV from data.

        Args:
            data: List of dictionaries
            headers: Optional list of headers to maintain column order

        Returns:
            CSV content as bytes
        """
        if not data:
            if headers:
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                return output.getvalue().encode("utf-8")
            return b""

        output = io.StringIO()
        fieldnames = headers if headers else list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue().encode("utf-8")

    @staticmethod
    def generate_xlsx(data: list[dict], sheet_name: str = "Items", headers: list[str] | None = None) -> bytes:
        """
        Generate XLSX from data.

        Args:
            data: List of dictionaries
            sheet_name: Name of the worksheet
            headers: Optional list of headers to maintain column order

        Returns:
            XLSX content as bytes
        """
        if not data:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            if headers:
                ws.append(headers)
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        df = pd.DataFrame(data)
        if headers:
            # Reorder columns according to headers and add missing ones as empty
            for col in headers:
                if col not in df.columns:
                    df[col] = None
            df = df[headers]
            
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return output.getvalue()

    @staticmethod
    def generate_json(data: list[dict]) -> bytes:
        """
        Generate JSON from data.

        Args:
            data: List of dictionaries

        Returns:
            JSON content as bytes
        """
        return json.dumps({"items": data}, indent=2, default=str).encode("utf-8")

    @staticmethod
    def generate_file(data: list[dict], file_format: str, headers: list[str] | None = None) -> bytes:
        """
        Generate file based on format.

        Args:
            data: List of dictionaries
            file_format: File format (csv, xlsx, json)
            headers: Optional list of headers to maintain column order

        Returns:
            File content as bytes
        """
        if file_format == FileFormat.CSV:
            return FileGenerator.generate_csv(data, headers=headers)
        elif file_format == FileFormat.XLSX:
            return FileGenerator.generate_xlsx(data, headers=headers)
        elif file_format == FileFormat.JSON:
            return FileGenerator.generate_json(data)
        else:
            raise ValueError(f"Unsupported format: {file_format}")


class ImportTemplate:
    """Generate import templates"""

    @staticmethod
    def get_template_columns() -> list[str]:
        """Get all valid columns for template"""
        return sorted(list(BulkImportValidator.VALID_COLUMNS))

    @staticmethod
    def get_template_data() -> list[dict]:
        """Get sample template data"""
        return [
            {
                "item_code": "ITEM001",
                "item_name": "Sample Item 1",
                "description": "This is a sample item",
                "item_group_name": "Default",
                "item_type": "stock",
                "status": "active",
                "uom": "Nos",
                "standard_rate": 100.00,
            }
        ]

    @staticmethod
    def get_template_csv() -> bytes:
        """Get CSV template"""
        return FileGenerator.generate_csv(ImportTemplate.get_template_data(), headers=ImportTemplate.get_template_columns())

    @staticmethod
    def get_template_xlsx() -> bytes:
        """Get XLSX template"""
        return FileGenerator.generate_xlsx(ImportTemplate.get_template_data(), headers=ImportTemplate.get_template_columns())

    @staticmethod
    def get_template_json() -> bytes:
        """Get JSON template"""
        return FileGenerator.generate_json(ImportTemplate.get_template_data())
